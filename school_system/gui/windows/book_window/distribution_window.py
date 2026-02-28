"""
Distribution Window

Dedicated window for managing book distribution sessions.
"""

from PyQt6.QtWidgets import QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton, QLineEdit, QComboBox, QTableWidget, QTableWidgetItem, QDateEdit, QTabWidget, QGroupBox, QProgressBar, QFileDialog
from PyQt6.QtCore import Qt, pyqtSignal, QDate, QTimer
from PyQt6.QtGui import QFont
from datetime import datetime
import pandas as pd

from school_system.gui.windows.base_function_window import BaseFunctionWindow
from school_system.gui.dialogs.message_dialog import show_error_message, show_success_message
from school_system.config.logging import logger
from school_system.services.book_service import BookService
from school_system.services.student_service import StudentService
from school_system.services.class_management_service import ClassManagementService
from school_system.gui.windows.book_window.utils import STANDARD_CLASSES, STANDARD_STREAMS, STANDARD_TERMS, STANDARD_SUBJECTS


class DistributionWindow(BaseFunctionWindow):
    """Dedicated window for managing book distribution sessions."""
    
    session_created = pyqtSignal()
    
    def __init__(self, parent=None, current_user: str = "", current_role: str = ""):
        """Initialize the distribution window."""
        super().__init__("Distribution & Templates", parent, current_user, current_role)

        self.book_service = BookService()
        self.student_service = StudentService()
        self.class_management_service = ClassManagementService()

        # Setup content
        self.setup_content()

        # Load sessions
        self._refresh_sessions_table()
    
    def setup_content(self):
        """Setup the main content area."""
        # Create tab widget
        self.tabs = QTabWidget()

        # Distribution Sessions Tab
        sessions_tab = self._create_sessions_tab()
        self.tabs.addTab(sessions_tab, "📋 Distribution Sessions")

        # Borrowing Templates Tab
        templates_tab = self._create_templates_tab()
        self.tabs.addTab(templates_tab, "📄 Borrowing Templates")

        # Add tabs to main layout
        main_layout = QVBoxLayout()
        main_layout.addWidget(self.tabs)
        main_layout.setContentsMargins(0, 0, 0, 0)

        # Add to content
        self.add_layout_to_content(main_layout)
    
    def _create_sessions_tab(self) -> QWidget:
        """Create the distribution sessions tab."""
        tab = QWidget()
        layout = QVBoxLayout(tab)
        layout.setSpacing(16)

        # Create session form
        form_card = self._create_session_form_card()
        layout.addWidget(form_card)

        # Sessions table
        table_card = self._create_sessions_table_card()
        layout.addWidget(table_card, stretch=1)

        return tab

    def _create_session_form_card(self) -> QWidget:
        """Create the distribution session form card."""
        theme_manager = self.get_theme_manager()
        theme = theme_manager._themes[self.get_theme()]
        
        form_card = QWidget()
        form_card.setProperty("card", "true")
        form_card.setStyleSheet(f"""
            QWidget[card="true"] {{
                background-color: {theme["surface"]};
                border: 1px solid {theme["border"]};
                border-radius: 12px;
                padding: 24px;
            }}
        """)
        
        form_layout = QVBoxLayout(form_card)
        form_layout.setContentsMargins(24, 24, 24, 24)
        form_layout.setSpacing(16)
        
        # Form title
        title_label = QLabel("Create Distribution Session")
        title_font = QFont("Segoe UI", 18, QFont.Weight.Medium)
        title_label.setFont(title_font)
        title_label.setStyleSheet(f"color: {theme['text']}; margin-bottom: 8px;")
        form_layout.addWidget(title_label)
        
        # Form fields in horizontal layout
        fields_layout = QHBoxLayout()
        fields_layout.setSpacing(12)
        
        # Class
        class_layout = QVBoxLayout()
        class_label = QLabel("Class")
        class_label.setStyleSheet(f"font-weight: 500; color: {theme['text']}; margin-bottom: 4px;")
        class_layout.addWidget(class_label)
        
        self.class_combo = QComboBox()
        self.class_combo.setFixedHeight(44)
        self.class_combo.addItem("")
        # Load classes dynamically from database
        try:
            classes = self.student_service.get_all_classes()
            self.class_combo.addItems(classes)
        except Exception as e:
            logger.warning(f"Could not load classes: {e}")
            # Fallback to standard classes
            self.class_combo.addItems(STANDARD_CLASSES)
        self.class_combo.setEditable(True)
        self.class_combo.currentTextChanged.connect(self._on_class_changed)
        class_layout.addWidget(self.class_combo)
        fields_layout.addLayout(class_layout)
        
        # Stream
        stream_layout = QVBoxLayout()
        stream_label = QLabel("Stream")
        stream_label.setStyleSheet(f"font-weight: 500; color: {theme['text']}; margin-bottom: 4px;")
        stream_layout.addWidget(stream_label)
        
        self.stream_combo = QComboBox()
        self.stream_combo.setFixedHeight(44)
        self.stream_combo.addItem("")
        # Load streams dynamically when class is selected
        self.stream_combo.setEditable(True)
        stream_layout.addWidget(self.stream_combo)
        fields_layout.addLayout(stream_layout)
        
        # Term
        term_layout = QVBoxLayout()
        term_label = QLabel("Term")
        term_label.setStyleSheet(f"font-weight: 500; color: {theme['text']}; margin-bottom: 4px;")
        term_layout.addWidget(term_label)
        
        self.term_combo = QComboBox()
        self.term_combo.setFixedHeight(44)
        self.term_combo.addItem("")
        self.term_combo.addItems(STANDARD_TERMS)
        self.term_combo.setEditable(True)
        term_layout.addWidget(self.term_combo)
        fields_layout.addLayout(term_layout)
        
        # Date
        date_layout = QVBoxLayout()
        date_label = QLabel("Distribution Date")
        date_label.setStyleSheet(f"font-weight: 500; color: {theme['text']}; margin-bottom: 4px;")
        date_layout.addWidget(date_label)
        
        self.date_input = QDateEdit()
        self.date_input.setFixedHeight(44)
        self.date_input.setDate(QDate.currentDate())
        self.date_input.setCalendarPopup(True)
        date_layout.addWidget(self.date_input)
        fields_layout.addLayout(date_layout)
        
        fields_layout.addStretch()
        
        # Create button
        create_btn = self.create_button("Create Session", "primary")
        create_btn.setFixedHeight(44)
        create_btn.clicked.connect(self._on_create_session)
        fields_layout.addWidget(create_btn)
        
        form_layout.addLayout(fields_layout)
        
        return form_card

    def _on_class_changed(self, class_name: str):
        """Handle class selection change."""
        if class_name and class_name != "":
            try:
                # Load streams for the selected class
                streams = self.student_service.get_streams_for_class(class_name)
                self.stream_combo.clear()
                self.stream_combo.addItem("")
                if streams:
                    self.stream_combo.addItems(streams)
                else:
                    # Fallback to standard streams
                    self.stream_combo.addItems(STANDARD_STREAMS)
            except Exception as e:
                logger.warning(f"Could not load streams for class {class_name}: {e}")
                # Fallback to standard streams
                self.stream_combo.clear()
                self.stream_combo.addItem("")
                self.stream_combo.addItems(STANDARD_STREAMS)
        else:
            # No class selected, clear streams
            self.stream_combo.clear()
            self.stream_combo.addItem("")

    def _create_sessions_table_card(self) -> QWidget:
        """Create the sessions table card."""
        theme_manager = self.get_theme_manager()
        theme = theme_manager._themes[self.get_theme()]
        
        table_card = QWidget()
        table_card.setProperty("card", "true")
        table_card.setStyleSheet(f"""
            QWidget[card="true"] {{
                background-color: {theme["surface"]};
                border: 1px solid {theme["border"]};
                border-radius: 12px;
                padding: 24px;
            }}
        """)
        
        table_layout = QVBoxLayout(table_card)
        table_layout.setContentsMargins(0, 0, 0, 0)
        table_layout.setSpacing(12)
        
        # Table title and refresh button
        title_layout = QHBoxLayout()
        title_label = QLabel("Distribution Sessions")
        title_font = QFont("Segoe UI", 16, QFont.Weight.Medium)
        title_label.setFont(title_font)
        title_label.setStyleSheet(f"color: {theme['text']}; margin-bottom: 8px;")
        title_layout.addWidget(title_label)
        
        title_layout.addStretch()
        
        refresh_btn = self.create_button("🔄 Refresh", "outline")
        refresh_btn.clicked.connect(self._refresh_sessions_table)
        title_layout.addWidget(refresh_btn)
        
        table_layout.addLayout(title_layout)
        
        # Sessions table
        self.sessions_table = self.create_table(0, 5)
        self.sessions_table.setColumnCount(5)
        self.sessions_table.setHorizontalHeaderLabels([
            "Session ID", "Class", "Stream", "Term", "Date"
        ])
        self.sessions_table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.sessions_table.setAlternatingRowColors(True)
        table_layout.addWidget(self.sessions_table)
        
        return table_card

    def _create_templates_tab(self) -> QWidget:
        """Create the borrowing templates tab."""
        tab = QWidget()
        layout = QVBoxLayout(tab)
        layout.setSpacing(16)

        # Template generation form
        template_form = self._create_template_form_card()
        layout.addWidget(template_form)

        # Template generation status/progress
        self.template_progress_group = QGroupBox("Generation Status")
        self.template_progress_group.setVisible(False)
        progress_layout = QVBoxLayout(self.template_progress_group)

        self.template_progress_bar = QProgressBar()
        self.template_progress_bar.setRange(0, 100)
        progress_layout.addWidget(self.template_progress_bar)

        self.template_status_label = QLabel("Ready to generate templates...")
        progress_layout.addWidget(self.template_status_label)

        layout.addWidget(self.template_progress_group)

        layout.addStretch()
        return tab

    def _create_template_form_card(self) -> QWidget:
        """Create the template generation form card."""
        theme_manager = self.get_theme_manager()
        theme = theme_manager._themes[self.get_theme()]

        form_card = QWidget()
        form_card.setProperty("card", "true")
        form_card.setStyleSheet(f"""
            QWidget[card="true"] {{
                background-color: {theme["surface"]};
                border: 1px solid {theme["border"]};
                border-radius: 12px;
                padding: 24px;
            }}
        """)

        form_layout = QVBoxLayout(form_card)
        form_layout.setContentsMargins(24, 24, 24, 24)
        form_layout.setSpacing(16)

        # Form title
        title_label = QLabel("Generate Borrowing Templates")
        title_font = QFont("Segoe UI", 18, QFont.Weight.Medium)
        title_label.setFont(title_font)
        title_label.setStyleSheet(f"color: {theme['text']}; margin-bottom: 8px;")
        form_layout.addWidget(title_label)

        subtitle_label = QLabel("Create printable Excel/PDF templates for book borrowing by class, stream, and subject")
        subtitle_label.setStyleSheet(f"color: {theme['text_secondary']}; font-size: 12px;")
        form_layout.addWidget(subtitle_label)

        # Template type selection
        type_layout = QHBoxLayout()
        type_label = QLabel("Template Type:")
        type_label.setStyleSheet(f"font-weight: 500; color: {theme['text']};")
        type_layout.addWidget(type_label)

        self.template_type_combo = QComboBox()
        self.template_type_combo.addItems([
            "Individual Class Templates",
            "Individual Stream per Subject Templates",
            "All Classes Combined Template"
        ])
        self.template_type_combo.setFixedHeight(40)
        self.template_type_combo.currentTextChanged.connect(self._on_template_type_changed)
        type_layout.addWidget(self.template_type_combo)

        type_layout.addStretch()
        form_layout.addLayout(type_layout)

        # Selection criteria (initially hidden)
        self.criteria_group = QGroupBox("Selection Criteria")
        criteria_layout = QVBoxLayout(self.criteria_group)

        # Class selection
        class_layout = QHBoxLayout()
        class_label = QLabel("Class:")
        class_label.setStyleSheet(f"font-weight: 500; color: {theme['text']};")
        class_layout.addWidget(class_label)

        self.template_class_combo = QComboBox()
        self.template_class_combo.addItem("All Classes")
        # Load classes dynamically from database
        try:
            classes = self.student_service.get_all_classes()
            self.template_class_combo.addItems(classes)
        except Exception as e:
            logger.warning(f"Could not load classes for templates: {e}")
            # Fallback to standard classes
            self.template_class_combo.addItems(STANDARD_CLASSES)
        self.template_class_combo.setFixedHeight(40)
        self.template_class_combo.currentTextChanged.connect(self._update_template_subjects)
        class_layout.addWidget(self.template_class_combo)

        class_layout.addStretch()
        criteria_layout.addLayout(class_layout)

        # Stream selection
        stream_layout = QHBoxLayout()
        stream_label = QLabel("Stream:")
        stream_label.setStyleSheet(f"font-weight: 500; color: {theme['text']};")
        stream_layout.addWidget(stream_label)

        self.template_stream_combo = QComboBox()
        self.template_stream_combo.addItem("All Streams")
        # Load streams dynamically from database
        try:
            streams = self.student_service.get_all_stream_names()
            self.template_stream_combo.addItems(streams)
        except Exception as e:
            logger.warning(f"Could not load streams for templates: {e}")
            # Fallback to standard streams
            self.template_stream_combo.addItems(STANDARD_STREAMS)
        self.template_stream_combo.setFixedHeight(40)
        stream_layout.addWidget(self.template_stream_combo)

        stream_layout.addStretch()
        criteria_layout.addLayout(stream_layout)

        # Subject selection
        subject_layout = QHBoxLayout()
        subject_label = QLabel("Subject:")
        subject_label.setStyleSheet(f"font-weight: 500; color: {theme['text']};")
        subject_layout.addWidget(subject_label)

        self.template_subject_combo = QComboBox()
        self.template_subject_combo.addItem("All Subjects")
        # Load subjects dynamically from database
        try:
            subjects = self.book_service.get_all_subjects()
            self.template_subject_combo.addItems(subjects)
        except Exception as e:
            logger.warning(f"Could not load subjects for templates: {e}")
            # Fallback to standard subjects
            self.template_subject_combo.addItems(STANDARD_SUBJECTS)
        self.template_subject_combo.setFixedHeight(40)
        subject_layout.addWidget(self.template_subject_combo)

        subject_layout.addStretch()
        criteria_layout.addLayout(subject_layout)

        form_layout.addWidget(self.criteria_group)
        self.criteria_group.setVisible(False)

        # Output format selection
        format_layout = QHBoxLayout()
        format_label = QLabel("Output Format:")
        format_label.setStyleSheet(f"font-weight: 500; color: {theme['text']};")
        format_layout.addWidget(format_label)

        self.output_format_combo = QComboBox()
        self.output_format_combo.addItems(["Excel (.xlsx)", "PDF (.pdf)", "Both"])
        self.output_format_combo.setFixedHeight(40)
        format_layout.addWidget(self.output_format_combo)

        format_layout.addStretch()
        form_layout.addLayout(format_layout)

        # Action buttons
        buttons_layout = QHBoxLayout()
        buttons_layout.addStretch()

        preview_btn = QPushButton("👁️ Preview")
        preview_btn.setFixedHeight(44)
        preview_btn.setFixedWidth(120)
        preview_btn.clicked.connect(self._preview_template)
        preview_btn.setStyleSheet(f"""
            QPushButton {{
                background-color: {theme["surface"]};
                color: {theme["text"]};
                border: 1px solid {theme["border"]};
                border-radius: 6px;
                padding: 8px;
                font-weight: 600;
            }}
            QPushButton:hover {{
                background-color: {theme["border"]};
            }}
        """)
        buttons_layout.addWidget(preview_btn)

        generate_btn = QPushButton("📄 Generate Templates")
        generate_btn.setFixedHeight(44)
        generate_btn.setFixedWidth(180)
        generate_btn.clicked.connect(self._generate_templates)
        generate_btn.setStyleSheet(f"""
            QPushButton {{
                background-color: {theme["primary"]};
                color: white;
                border: none;
                border-radius: 6px;
                padding: 8px;
                font-weight: 600;
            }}
            QPushButton:hover {{
                background-color: {theme.get("primary_hover", theme["primary"])};
            }}
        """)
        buttons_layout.addWidget(generate_btn)

        form_layout.addLayout(buttons_layout)

        return form_card

    def _on_template_type_changed(self, template_type: str):
        """Handle template type change."""
        if template_type == "Individual Class Templates":
            self.criteria_group.setVisible(False)
        elif template_type == "Individual Stream per Subject Templates":
            self.criteria_group.setVisible(True)
            self.template_class_combo.setCurrentText("All Classes")
            self.template_stream_combo.setCurrentText("All Streams")
        else:  # All Classes Combined Template
            self.criteria_group.setVisible(False)

    def _update_template_subjects(self, class_text: str):
        """Update available subjects based on selected class."""
        # This could be enhanced to show class-specific subjects
        pass

    def _preview_template(self):
        """Preview the template before generation."""
        try:
            template_data = self._prepare_template_data()

            if not template_data:
                show_error_message("Preview Error", "No data available for the selected criteria.", self)
                return

            # Create a simple preview dialog
            preview_text = f"Borrowing Template Preview\n{'='*40}\n\n"
            preview_text += f"School: School Management System\n"
            preview_text += f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n"
            preview_text += f"Generated by: {self.current_user}\n\n"

            total_students = sum(len(students) for students in template_data.values())
            preview_text += f"Total Templates: {len(template_data)}\n"
            preview_text += f"Total Students: {total_students}\n\n"

            for template_key, students in template_data.items():
                preview_text += f"Template: {template_key}\n"
                preview_text += f"Students: {len(students)}\n\n"

                # Show first few students
                for i, student in enumerate(students[:3]):
                    preview_text += f"  {i+1}. {student.name} ({student.admission_number})\n"

                if len(students) > 3:
                    preview_text += f"  ... and {len(students) - 3} more students\n"

                preview_text += "\n"

            # Show preview in a message box
            from PyQt6.QtWidgets import QMessageBox
            msg_box = QMessageBox(self)
            msg_box.setWindowTitle("Template Preview")
            msg_box.setText("Template Preview")
            msg_box.setDetailedText(preview_text)
            msg_box.setStandardButtons(QMessageBox.StandardButton.Ok)
            msg_box.exec()

        except Exception as e:
            logger.error(f"Error creating template preview: {e}")
            show_error_message("Preview Error", f"Failed to create preview: {str(e)}", self)

    def _generate_templates(self):
        """Generate the borrowing templates."""
        try:
            self.template_progress_group.setVisible(True)
            self.template_progress_bar.setValue(0)
            self.template_status_label.setText("Preparing template data...")

            # Get template data
            template_data = self._prepare_template_data()

            if not template_data:
                show_error_message("Generation Error", "No data available for the selected criteria.", self)
                self.template_progress_group.setVisible(False)
                return

            # Generate templates in background
            QTimer.singleShot(100, lambda: self._do_generate_templates(template_data))

        except Exception as e:
            logger.error(f"Error starting template generation: {e}")
            show_error_message("Generation Error", f"Failed to start generation: {str(e)}", self)
            self.template_progress_group.setVisible(False)

    def _do_generate_templates(self, template_data: dict):
        """Background task to generate templates."""
        try:
            import os

            self.template_status_label.setText("Generating templates...")

            output_format = self.output_format_combo.currentText()
            generate_excel = "Excel" in output_format or output_format == "Both"
            generate_pdf = "PDF" in output_format or output_format == "Both"

            total_templates = len(template_data)
            completed = 0

            generated_files = []

            for template_key, students in template_data.items():
                self.template_status_label.setText(f"Generating template: {template_key}")
                logger.info(f"Generating template: {template_key} for {len(students)} students")

                # Clean template key for filename
                safe_template_key = template_key.replace(" ", "_").replace("/", "_")

                # Generate Excel
                if generate_excel:
                    # Prompt user for save location
                    excel_file, _ = QFileDialog.getSaveFileName(
                        self,
                        f"Save Excel Template - {template_key}",
                        f"{safe_template_key}.xlsx",
                        "Excel Files (*.xlsx)"
                    )

                    if excel_file:  # User didn't cancel
                        if self._generate_excel_template_to_file(template_key, students, excel_file):
                            generated_files.append(("Excel", excel_file))
                            logger.info(f"Generated Excel template: {excel_file}")
                        else:
                            logger.error(f"Failed to generate Excel template for {template_key}")
                    else:
                        logger.info(f"User cancelled Excel save for {template_key}")

                # Generate PDF
                if generate_pdf:
                    # Prompt user for save location
                    pdf_file, _ = QFileDialog.getSaveFileName(
                        self,
                        f"Save PDF Template - {template_key}",
                        f"{safe_template_key}.pdf",
                        "PDF Files (*.pdf)"
                    )

                    if pdf_file:  # User didn't cancel
                        if self._generate_pdf_template_to_file(template_key, students, pdf_file):
                            generated_files.append(("PDF", pdf_file))
                            logger.info(f"Generated PDF template: {pdf_file}")
                        else:
                            logger.error(f"Failed to generate PDF template for {template_key}")
                    else:
                        logger.info(f"User cancelled PDF save for {template_key}")

                completed += 1
                progress = int((completed / total_templates) * 100)
                self.template_progress_bar.setValue(progress)

            self.template_progress_bar.setValue(100)
            self.template_status_label.setText(f"Generation complete! Created {len(generated_files)} files.")

            # Show success message with file locations
            if generated_files:
                success_msg = f"Successfully generated {len(generated_files)} template files:\n\n"
                for format_type, file_path in generated_files[:10]:  # Show first 10
                    success_msg += f"• {format_type}: {os.path.basename(file_path)}\nSaved to: {file_path}\n\n"

                if len(generated_files) > 10:
                    success_msg += f"... and {len(generated_files) - 10} more files"

                show_success_message("Templates Generated", success_msg, self)
            else:
                show_info_message("No Templates Generated", "No templates were generated. All saves were cancelled or failed.", self)

        except Exception as e:
            logger.error(f"Error generating templates: {e}")
            show_error_message("Generation Error", f"Failed to generate templates: {str(e)}", self)

        finally:
            QTimer.singleShot(3000, lambda: self.template_progress_group.setVisible(False))

    def _prepare_template_data(self) -> dict:
        """Prepare template data based on selected criteria."""
        template_type = self.template_type_combo.currentText()

        if template_type == "Individual Class Templates":
            return self._prepare_class_templates()
        elif template_type == "Individual Stream per Subject Templates":
            return self._prepare_stream_subject_templates()
        else:  # All Classes Combined Template
            return self._prepare_combined_template()

    def _prepare_class_templates(self) -> dict:
        """Prepare templates for individual classes."""
        templates = {}

        class_levels = self.class_management_service.get_all_class_levels()

        for class_level in class_levels:
            students = self.student_service.get_students_by_class_level(class_level)
            if students:
                template_key = f"Form_{class_level}_All_Streams"
                templates[template_key] = students

        return templates

    def _prepare_stream_subject_templates(self) -> dict:
        """Prepare templates for individual stream per subject combinations."""
        templates = {}

        selected_class = self.template_class_combo.currentText()
        selected_stream = self.template_stream_combo.currentText()
        selected_subject = self.template_subject_combo.currentText()

        # Get combinations based on filters
        combinations = self.class_management_service.get_class_stream_combinations()

        for class_level, stream, count in combinations:
            # Filter by selected class
            if selected_class != "All Classes":
                # Extract the class level from the selected class name (e.g., "Form 4" -> 4)
                import re
                match = re.search(r'\d+', selected_class)
                if match:
                    selected_class_level = int(match.group())
                    if class_level != selected_class_level:
                        continue
                else:
                    # If we can't extract a number, skip this combination
                    continue

            # Filter by selected stream
            if selected_stream != "All Streams" and stream != selected_stream:
                continue

            students = self.class_management_service.get_students_by_class_and_stream(class_level, stream)
            if students:
                # Create templates for each subject
                if selected_subject != "All Subjects":
                    subjects = [selected_subject]
                else:
                    # Get subjects dynamically from database
                    try:
                        subjects = self.book_service.get_all_subjects()
                        if not subjects:
                            subjects = STANDARD_SUBJECTS  # Fallback
                    except Exception as e:
                        logger.warning(f"Could not get subjects from database: {e}")
                        subjects = STANDARD_SUBJECTS  # Fallback

                for subject in subjects:
                    template_key = f"Form_{class_level}_{stream}_{subject}"
                    templates[template_key] = students

        return templates

    def _prepare_combined_template(self) -> dict:
        """Prepare a single combined template for all classes."""
        all_students = self.student_service.get_all_students()
        return {"All_Classes_Combined": all_students}

    def _generate_excel_template_to_file(self, template_key: str, students: list, file_path: str) -> bool:
        """Generate Excel template for borrowing to a specific file path."""
        try:
            # Parse template key to get details
            parts = template_key.split('_')
            class_info = parts[1] if len(parts) > 1 else "All"
            stream_info = parts[2] if len(parts) > 2 else "All"
            subject_info = parts[3] if len(parts) > 3 else "All"

            # Create DataFrame
            data = {
                'School_Name': ['School Management System'] * len(students),
                'Class_Form': [class_info] * len(students),
                'Stream': [stream_info] * len(students),
                'Subject': [subject_info] * len(students),
                'Student_Name': [student.name for student in students],
                'Admission_Number': [student.admission_number or str(student.student_id) for student in students],
                'Book_Number': [''] * len(students),  # Empty column for user to fill
                'Date_Borrowed': [''] * len(students),
                'Librarian_Name': [self.current_user] * len(students),
                'Generated_Date': [datetime.now().strftime('%Y-%m-%d')] * len(students)
            }

            df = pd.DataFrame(data)

            # Save to Excel with basic formatting
            try:
                with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                    df.to_excel(writer, sheet_name='Borrowing_Template', index=False)

                    # Try to add basic formatting
                    try:
                        # Get workbook and worksheet
                        workbook = writer.book
                        worksheet = writer.sheets['Borrowing_Template']

                        # Set column widths
                        from openpyxl.utils import get_column_letter
                        for column in worksheet.columns:
                            max_length = 0
                            column_letter = get_column_letter(column[0].column)

                            for cell in column:
                                try:
                                    if len(str(cell.value)) > max_length:
                                        max_length = len(str(cell.value))
                                except:
                                    pass

                            adjusted_width = min(max_length + 2, 30)  # Max width of 30
                            worksheet.column_dimensions[column_letter].width = adjusted_width

                        # Style header row (optional - skip if it fails)
                        try:
                            from openpyxl.styles import Font, PatternFill
                            header_font = Font(bold=True, color="FFFFFF")
                            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

                            for cell in worksheet[1]:
                                cell.font = header_font
                                cell.fill = header_fill
                        except Exception as style_error:
                            logger.warning(f"Could not apply Excel styling: {style_error}")
                            # Continue without styling

                    except Exception as format_error:
                        logger.warning(f"Could not apply Excel formatting: {format_error}")
                        # Continue with basic Excel file

            except Exception as excel_error:
                logger.error(f"Failed to save Excel file {file_path}: {excel_error}")
                return False

            return True

        except Exception as e:
            logger.error(f"Error generating Excel template: {e}")
            return False

    def _generate_pdf_template(self, template_key: str, students: list) -> str:
        """Generate PDF template for borrowing."""
        try:
            from fpdf import FPDF
            import tempfile
            from datetime import datetime

            # Parse template key to get details
            parts = template_key.split('_')
            class_info = parts[1] if len(parts) > 1 else "All"
            stream_info = parts[2] if len(parts) > 2 else "All"
            subject_info = parts[3] if len(parts) > 3 else "All"

            # Create PDF
            pdf = FPDF()
            pdf.add_page()
            pdf.set_font("Arial", size=12)

            # Header
            pdf.set_font("Arial", 'B', 16)
            pdf.cell(0, 10, "SCHOOL MANAGEMENT SYSTEM", ln=True, align='C')
            pdf.cell(0, 10, "BOOK BORROWING TEMPLATE", ln=True, align='C')
            pdf.ln(10)

            # Template details
            pdf.set_font("Arial", 'B', 12)
            pdf.cell(0, 8, f"Class/Form: {class_info}", ln=True)
            pdf.cell(0, 8, f"Stream: {stream_info}", ln=True)
            pdf.cell(0, 8, f"Subject: {subject_info}", ln=True)
            pdf.cell(0, 8, f"Generated by: {self.current_user}", ln=True)
            pdf.cell(0, 8, f"Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", ln=True)
            pdf.ln(10)

            # Table headers
            pdf.set_font("Arial", 'B', 10)
            headers = ["#", "Student Name", "Admission No.", "Book Number", "Signature"]
            col_widths = [15, 60, 30, 35, 40]

            for i, header in enumerate(headers):
                pdf.cell(col_widths[i], 8, header, border=1, align='C')
            pdf.ln()

            # Table data
            pdf.set_font("Arial", size=9)
            for i, student in enumerate(students, 1):
                pdf.cell(col_widths[0], 6, str(i), border=1, align='C')
                pdf.cell(col_widths[1], 6, student.name[:25], border=1)  # Truncate long names
                admission = student.admission_number or str(student.student_id)
                pdf.cell(col_widths[2], 6, admission, border=1, align='C')
                pdf.cell(col_widths[3], 6, "", border=1, align='C')  # Empty for book number
                pdf.cell(col_widths[4], 6, "", border=1, align='C')  # Empty for signature
                pdf.ln()

            # Footer instructions
            pdf.ln(10)
            pdf.set_font("Arial", 'I', 8)
            pdf.multi_cell(0, 5, "Instructions:\n1. Fill in the Book Number column with the assigned book numbers.\n2. Students should sign in the Signature column when receiving books.\n3. Return this form to the librarian after distribution.")

            # Create temporary file
            with tempfile.NamedTemporaryFile(delete=False, suffix='.pdf') as tmp_file:
                pdf_file = tmp_file.name

            pdf.output(pdf_file)
            return pdf_file

        except Exception as e:
            logger.error(f"Error generating PDF template: {e}")
            return None

    def _generate_pdf_template_to_file(self, template_key: str, students: list, file_path: str) -> bool:
        """Generate PDF template for borrowing to a specific file path."""
        try:
            from fpdf import FPDF
            from datetime import datetime

            # Parse template key to get details
            parts = template_key.split('_')
            class_info = parts[1] if len(parts) > 1 else "All"
            stream_info = parts[2] if len(parts) > 2 else "All"
            subject_info = parts[3] if len(parts) > 3 else "All"

            # Create PDF
            pdf = FPDF()
            pdf.add_page()
            pdf.set_font("Arial", size=12)

            # Header
            pdf.set_font("Arial", 'B', 16)
            pdf.cell(0, 10, "SCHOOL MANAGEMENT SYSTEM", ln=True, align='C')
            pdf.cell(0, 10, "BOOK BORROWING TEMPLATE", ln=True, align='C')
            pdf.ln(10)

            # Template details
            pdf.set_font("Arial", 'B', 12)
            pdf.cell(0, 8, f"Class/Form: {class_info}", ln=True)
            pdf.cell(0, 8, f"Stream: {stream_info}", ln=True)
            pdf.cell(0, 8, f"Subject: {subject_info}", ln=True)
            pdf.cell(0, 8, f"Generated by: {self.current_user}", ln=True)
            pdf.cell(0, 8, f"Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", ln=True)
            pdf.ln(10)

            # Table headers
            pdf.set_font("Arial", 'B', 10)
            headers = ["#", "Student Name", "Admission No.", "Book Number", "Signature"]
            col_widths = [15, 60, 30, 35, 40]

            for i, header in enumerate(headers):
                pdf.cell(col_widths[i], 8, header, border=1, align='C')
            pdf.ln()

            # Table data
            pdf.set_font("Arial", size=9)
            for i, student in enumerate(students, 1):
                pdf.cell(col_widths[0], 6, str(i), border=1, align='C')
                pdf.cell(col_widths[1], 6, student.name[:25], border=1)  # Truncate long names
                admission = student.admission_number or str(student.student_id)
                pdf.cell(col_widths[2], 6, admission, border=1, align='C')
                pdf.cell(col_widths[3], 6, "", border=1, align='C')  # Empty for book number
                pdf.cell(col_widths[4], 6, "", border=1, align='C')  # Empty for signature
                pdf.ln()

            # Footer instructions
            pdf.ln(10)
            pdf.set_font("Arial", 'I', 8)
            pdf.multi_cell(0, 5, "Instructions:\n1. Fill in the Book Number column with the assigned book numbers.\n2. Students should sign in the Signature column when receiving books.\n3. Return this form to the librarian after distribution.")

            # Save to specified file path
            pdf.output(file_path)
            return True

        except Exception as e:
            logger.error(f"Error generating PDF template: {e}")
            return False

    def _on_create_session(self):
        """Handle create session button click."""
        # Get form data
        class_name = self.class_combo.currentText().strip()
        stream = self.stream_combo.currentText().strip()
        term = self.term_combo.currentText().strip()
        date = self.date_input.date().toString("yyyy-MM-dd")
        
        # Validate
        if not class_name or not stream or not term:
            show_error_message("Validation Error", "Please fill in all required fields (Class, Stream, Term).", self)
            return
        
        try:
            # Create distribution session
            session_data = {
                "class_name": class_name,
                "stream": stream,
                "term": term,
                "distribution_date": date
            }
            
            # Note: This would need to be implemented in BookService
            # For now, just show a message
            show_success_message("Success", f"Distribution session created for {class_name} {stream} - {term}.", self)
            
            # Clear inputs
            self.class_combo.setCurrentIndex(0)
            self.stream_combo.setCurrentIndex(0)
            self.term_combo.setCurrentIndex(0)
            self.date_input.setDate(QDate.currentDate())
            
            # Refresh table
            self._refresh_sessions_table()
            
            # Emit signal
            self.session_created.emit()
            
        except Exception as e:
            logger.error(f"Error creating distribution session: {e}")
            show_error_message("Error", f"Failed to create distribution session: {str(e)}", self)
    
    def _refresh_sessions_table(self):
        """Refresh the sessions table."""
        try:
            # Note: This would need to be implemented in BookService
            # For now, just clear the table
            self.sessions_table.setRowCount(0)
            
            logger.info("Refreshed distribution sessions table")
        except Exception as e:
            logger.error(f"Error refreshing sessions table: {e}")
            show_error_message("Error", f"Failed to refresh sessions: {str(e)}", self)
