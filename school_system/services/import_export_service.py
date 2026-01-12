"""
Import/Export service for handling data import and export operations.
"""

import csv
import json
import openpyxl
from typing import List, Dict, Tuple
from school_system.config.logging import logger
from school_system.config.settings import Settings
from school_system.core.exceptions import FileOperationException
from school_system.core.utils import ValidationUtils
from fpdf import FPDF


class ImportExportService:
    """Service for handling data import and export operations."""

    def export_to_csv(self, data: List[Dict], filename: str) -> bool:
        """
        Export data to a CSV file.

        Args:
            data: The data to export.
            filename: The name of the CSV file.

        Returns:
            True if the export was successful, otherwise False.
        """
        logger.info(f"Exporting data to CSV file: {filename}")
        ValidationUtils.validate_input(filename, "Filename cannot be empty")
        
        try:
            with open(filename, mode='w', newline='', encoding='utf-8') as file:
                writer = csv.DictWriter(file, fieldnames=data[0].keys())
                writer.writeheader()
                writer.writerows(data)
            logger.info(f"Data exported successfully to {filename}")
            return True
        except Exception as e:
            logger.error(f"Error exporting to CSV: {e}")
            return False

    def import_from_csv(self, filename: str) -> List[Dict]:
        """
        Import data from a CSV file.

        Args:
            filename: The name of the CSV file.

        Returns:
            The imported data as a list of dictionaries.
        """
        try:
            with open(filename, mode='r', encoding='utf-8') as file:
                reader = csv.DictReader(file)
                return list(reader)
        except Exception as e:
            print(f"Error importing from CSV: {e}")
            return []

    def export_to_json(self, data: List[Dict], filename: str) -> bool:
        """
        Export data to a JSON file.

        Args:
            data: The data to export.
            filename: The name of the JSON file.

        Returns:
            True if the export was successful, otherwise False.
        """
        try:
            with open(filename, mode='w', encoding='utf-8') as file:
                json.dump(data, file, indent=4)
            return True
        except Exception as e:
            print(f"Error exporting to JSON: {e}")
            return False

    def import_from_json(self, filename: str) -> List[Dict]:
        """
        Import data from a JSON file.

        Args:
            filename: The name of the JSON file.

        Returns:
            The imported data as a list of dictionaries.
        """
        try:
            with open(filename, mode='r', encoding='utf-8') as file:
                return json.load(file)
        except Exception as e:
            print(f"Error importing from JSON: {e}")
            return []

    def export_to_excel(self, data: List[Dict], filename: str) -> bool:
        """
        Export data to an Excel file.

        Args:
            data: The data to export.
            filename: The name of the Excel file.

        Returns:
            True if the export was successful, otherwise False.
        """
        logger.info(f"Exporting data to Excel file: {filename}")
        ValidationUtils.validate_input(filename, "Filename cannot be empty")
        
        try:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            
            if data:
                headers = list(data[0].keys())
                sheet.append(headers)
                
                for row in data:
                    sheet.append([row.get(header, '') for header in headers])
            
            workbook.save(filename)
            logger.info(f"Data exported successfully to {filename}")
            return True
        except Exception as e:
            logger.error(f"Error exporting to Excel: {e}")
            return False

    def import_from_excel(self, filename: str) -> List[Dict]:
        """
        Import data from an Excel file.

        Args:
            filename: The name of the Excel file.

        Returns:
            The imported data as a list of dictionaries.
        """
        try:
            workbook = openpyxl.load_workbook(filename)
            sheet = workbook.active
            
            headers = [cell.value for cell in sheet[1]]
            data = []
            
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if row:
                    data.append(dict(zip(headers, row)))
            
            return data
        except Exception as e:
            print(f"Error importing from Excel: {e}")
            return []
    
    def import_from_excel_with_validation(self, filename: str, required_columns: List[str]) -> Tuple[bool, List[Dict], str]:
        """
        Import data from Excel with validation.
        
        Args:
            filename: The name of the Excel file.
            required_columns: List of required column names
            
        Returns:
            tuple: (success, data, error_message)
        """
        try:
            workbook = openpyxl.load_workbook(filename)
            sheet = workbook.active
            
            # Get headers from first row
            headers = [cell.value for cell in sheet[1]]
            
            # Validate required columns
            missing_columns = []
            for column in required_columns:
                if column not in headers:
                    missing_columns.append(column)
            
            if missing_columns:
                return False, [], f"Missing required columns: {', '.join(missing_columns)}"
            
            # Import data
            data = []
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if row:
                    data.append(dict(zip(headers, row)))
            
            return True, data, ""
            
        except Exception as e:
            logger.error(f"Error importing from Excel with validation: {e}")
            return False, [], f"Error importing file: {str(e)}"
    
    def generate_excel_template(self, filename: str, columns: List[str], sample_data: List[Dict] = None) -> bool:
        """
        Generate an Excel template file with proper headers.
        
        Args:
            filename: The name of the Excel file to create.
            columns: List of column headers.
            sample_data: Optional sample data to include.
            
        Returns:
            True if template generation was successful, False otherwise.
        """
        try:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            
            # Add headers
            sheet.append(columns)
            
            # Add sample data if provided
            if sample_data:
                for row in sample_data:
                    sheet.append([row.get(col, '') for col in columns])
            
            # Auto-size columns
            for column in sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2
                sheet.column_dimensions[column_letter].width = adjusted_width
            
            workbook.save(filename)
            logger.info(f"Excel template generated successfully: {filename}")
            return True
            
        except Exception as e:
            logger.error(f"Error generating Excel template: {e}")
            return False

    def export_to_pdf(self, data: List[Dict], filename: str) -> bool:
        """
        Export data to a PDF file.
        
        Args:
            data: The data to export.
            filename: The name of the PDF file.
            
        Returns:
            True if the export was successful, otherwise False.
        """
        logger.info(f"Exporting data to PDF file: {filename}")
        ValidationUtils.validate_input(filename, "Filename cannot be empty")
        
        try:
            pdf = FPDF()
            pdf.add_page()
            pdf.set_font("Arial", size=12)
            
            if data:
                # Add headers
                headers = list(data[0].keys())
                col_width = 40
                
                # Add headers
                for header in headers:
                    pdf.cell(col_width, 10, str(header), border=1)
                pdf.ln()
                
                # Add data rows
                for row in data:
                    for header in headers:
                        pdf.cell(col_width, 10, str(row.get(header, '')), border=1)
                    pdf.ln()
            
            pdf.output(filename)
            logger.info(f"Data exported successfully to {filename}")
            return True
        except Exception as e:
            logger.error(f"Error exporting to PDF: {e}")
            return False

    def import_from_pdf(self, filename: str) -> List[Dict]:
        """
        Import data from a PDF file.
        
        Args:
            filename: The name of the PDF file.
            
        Returns:
            The imported data as a list of dictionaries.
        """
        try:
            # Note: PDF import is complex and typically requires OCR or specific formatting.
            # This is a placeholder for a more advanced implementation.
            logger.warning("PDF import is not fully implemented. Consider using CSV or Excel for data import.")
            return []
        except Exception as e:
            logger.error(f"Error importing from PDF: {e}")
            return []