# library_logic.py

import sqlite3
from sqlite3 import Error as SQLiteError
from tkinter import messagebox
import pandas as pd
from datetime import datetime, date
import tkinter as tk
import qrcode
import fitz
from PIL import Image
import os
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from pdf2image import convert_from_bytes
import io
import re
import openpyxl
import winsound
import hashlib
import json
import logging



class LibraryLogic:
    def __init__(self, db_manager):
        """Initialize LibraryLogic with a database manager."""
        self.logger = logging.getLogger('LibraryLogic')
        self.db_manager = db_manager
        self.logger.info("Initializing LibraryLogic")
        try:
            self.users = self.db_manager.load_users()
            self.teachers_data = self.db_manager.load_teachers()
            self.borrowed_books = []
            self.logger.info("LibraryLogic initialized successfully")
        except SQLiteError as db_err:
            self.logger.error(f"Failed to load initial data: {db_err}")
            messagebox.showerror("Initialization Error", "Failed to load initial data.")
            self.users = {}
            self.teachers_data = {}
            self.borrowed_books = []

    def login(self, username, password):
        """Authenticate a user and return True if successful."""
        try:
            hashed_password = hashlib.sha256(password.encode()).hexdigest()
            user_info = self.users.get(username)
            if user_info and user_info["password"] == hashed_password:
                return True
            messagebox.showerror("Error", "Invalid username or password.")
            return False
        except TypeError as type_err:
            messagebox.showerror("Error", f"Invalid input type for login: {type_err}")
            return False

    def save_user(self, username, password, role):
        """Save a new user with hashed password."""
        try:
            if not all([username, password, role]):
                messagebox.showerror("Error", "Please fill in all fields.")
                return False
            hashed_password = hashlib.sha256(password.encode()).hexdigest()
            if self.db_manager.save_user(username, hashed_password, role):
                self.users[username] = {"password": hashed_password, "role": role}
                return True
            return False
        except TypeError as type_err:
            messagebox.showerror("Error", f"Invalid input type for user: {type_err}")
            return False
        except SQLiteError as db_err:
            messagebox.showerror("Database Error", f"Failed to save user: {db_err}")
            return False

    def load_students_data(self, file_path):
        """Load student data from an Excel file into the database."""
        try:
            df = pd.read_excel(file_path)
            for _, row in df.iterrows():
                student_id = str(row['Student_ID'])
                name = str(row['Name'])
                stream = str(row['Stream']) if 'Stream' in df.columns and pd.notna(row['Stream']) else None
                self.db_manager.save_student(student_id, name, stream)
            self.logger.info(f"Loaded {len(df)} students from {file_path}")
            return True
        except FileNotFoundError as fnf_err:
            messagebox.showerror("File Error", f"Excel file not found: {fnf_err}")
            return False
        except pd.errors.EmptyDataError as empty_err:
            messagebox.showerror("File Error", f"Excel file is empty: {empty_err}")
            return False
        except KeyError as key_err:
            messagebox.showerror("File Error", f"Missing required column (Student_ID or Name): {key_err}")
            return False
        except SQLiteError as db_err:
            messagebox.showerror("Database Error", f"Failed to load students: {db_err}")
            return False

    def load_books_from_excel(self, file_path):
        """Load books from an Excel file into the database."""
        try:
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active
            book_ids = [str(row[0].value) for row in sheet.iter_rows(min_row=2) if row[0].value]
            conditions = [str(row[1].value) if len(row) > 1 and row[1].value else "New" 
                         for row in sheet.iter_rows(min_row=2) if row[0].value]
            book_count = self.db_manager.load_books_from_excel(book_ids, conditions)
            if book_count > 0:
                messagebox.showinfo("Success", f"{book_count} books loaded successfully!")
                return True
            messagebox.showinfo("Info", "No books loaded from file.")
            return False
        except FileNotFoundError as fnf_err:
            messagebox.showerror("File Error", f"Excel file not found: {fnf_err}")
            return False
        except openpyxl.utils.exceptions.InvalidFileException as inv_err:
            messagebox.showerror("File Error", f"Invalid Excel file format: {inv_err}")
            return False

    def bulk_borrow_books(self, file_path=None, borrowings=None):
        """Handle bulk borrowing from an Excel file or a list."""
        if file_path:
            try:
                workbook = openpyxl.load_workbook(file_path)
                sheet = workbook.active
                borrowings = [
                    (str(row[0].value), str(row[1].value), str(row[2].value), 
                     str(row[3].value) if len(row) > 3 and row[3].value else "New")
                    for row in sheet.iter_rows(min_row=2) if row[0].value and row[1].value and row[2].value
                ]
            except Exception as e:
                messagebox.showerror("File Error", f"Failed to read Excel file: {e}")
                return {"success": [], "already_borrowed": [], "errors": []}

        if not borrowings:
            messagebox.showerror("Error", "No valid borrowing data provided.")
            return {"success": [], "already_borrowed": [], "errors": []}

        results = self.db_manager.bulk_borrow(borrowings)
        return results

    def bulk_return_books(self, file_path=None, returns=None):
        """Handle bulk returning from an Excel file or a list."""
        if file_path:
            try:
                workbook = openpyxl.load_workbook(file_path)
                sheet = workbook.active
                returns = [
                    (str(row[0].value), str(row[1].value), str(row[2].value), 
                     str(row[3].value) if len(row) > 3 and row[3].value else None)
                    for row in sheet.iter_rows(min_row=2) if row[0].value and row[1].value and row[2].value
                ]
            except Exception as e:
                messagebox.showerror("File Error", f"Failed to read Excel file: {e}")
                return 0, []

        if not returns:
            messagebox.showerror("Error", "No valid return data provided.")
            return 0, []

        successes, failures = self.db_manager.bulk_return(returns)
        if successes > 0:
            messagebox.showinfo("Success", f"Successfully returned {successes} books. Failures: {len(failures)}")
        if failures:
            messagebox.showwarning("Failures", f"Failed operations:\n" + 
                                  "\n".join(f"User {u}, Book {b}: {r}" for u, b, r in failures))
        return successes, failures

    def save_student(self, student_id, name, stream):
        """Save a new student with mandatory stream."""
        try:
            return self.db_manager.save_student(student_id, name, stream)
        except Exception:
            return False

            
    def add_teacher(self, teacher_id, teacher_name):
        """Add a new teacher."""
        try:
            if not all([teacher_id, teacher_name]):
                messagebox.showerror("Warning", "Please enter teacher name and TSC number.")
                return False
            if self.db_manager.save_teacher(teacher_id, teacher_name):
                self.teachers_data[teacher_id] = {"teacher_name": teacher_name}
                return True
            return False
        except TypeError as type_err:
            messagebox.showerror("Error", f"Invalid input type for teacher: {type_err}")
            return False
        except SQLiteError as db_err:
            messagebox.showerror("Database Error", f"Failed to add teacher: {db_err}")
            return False

    def add_new_book(self, book_id):
        """Add a new book."""
        try:
            if not book_id:
                messagebox.showerror("Error", "Please enter Book ID.")
                return False
            return self.db_manager.save_book(book_id)
        except TypeError as type_err:
            messagebox.showerror("Error", f"Invalid input type for book: {type_err}")
            return False
        except SQLiteError as db_err:
            messagebox.showerror("Database Error", f"Failed to add book: {db_err}")
            return False

    def add_book_with_tags(self, book_id, tags=None):
        """Add a new book with optional tags."""
        try:
            if not book_id:
                messagebox.showerror("Error", "Please enter Book ID.")
                return False
            if tags:
                tags = [tag.strip() for tag in tags.split(",") if tag.strip()]
            return self.db_manager.save_book(book_id, tags)
        except TypeError as type_err:
            messagebox.showerror("Error", f"Invalid input type for book: {type_err}")
            return False
        except SQLiteError as db_err:
            messagebox.showerror("Database Error", f"Failed to add book: {db_err}")
            return False

    def tag_book(self, book_id, tags):
        """Add tags to an existing book."""
        try:
            if not book_id:
                messagebox.showerror("Error", "Please enter Book ID.")
                return False
            tags = [tag.strip() for tag in tags.split(",") if tag.strip()]
            conn = self.db_manager._create_connection()
            if not conn:
                return False
            cursor = conn.cursor()
            try:
                for tag in tags:
                    cursor.execute("INSERT OR IGNORE INTO book_tags (book_id, tag) VALUES (?, ?)", (book_id, tag))
                conn.commit()
                return True
            except SQLiteError as db_err:
                conn.rollback()
                messagebox.showerror("Database Error", f"Tag addition failed: {db_err}")
                return False
            finally:
                self.db_manager._close_connection(conn)
        except TypeError as type_err:
            messagebox.showerror("Error", f"Invalid input type for tags: {type_err}")
            return False

    def borrow_book_student(self, student_id, book_ids, reminder_days=None, condition="New"):
        """Handle a student borrowing books."""
        try:
            if not all([student_id, book_ids, condition]):
                messagebox.showerror("Error", "Please fill in all fields, including condition.")
                return False
            if condition not in ["New", "Good", "Damaged"]:
                messagebox.showerror("Error", "Invalid condition selected.")
                return False
            success, feedback = self.db_manager.borrow_book_student(student_id, book_ids, date.today(), reminder_days, condition)
            if success:
                messagebox.showinfo("Borrowing Result", feedback)
            else:
                messagebox.showerror("Error", feedback)
            return success
        except ValueError as val_err:
            messagebox.showerror("Error", f"Invalid reminder days: {val_err}")
            return False

    def borrow_book_teacher(self, teacher_id, book_id, condition="New"):
        """Handle a teacher borrowing a book."""
        try:
            if not all([teacher_id, book_id, condition]):
                messagebox.showerror("Error", "Please fill in all fields, including condition.")
                return False
            if condition not in ["New", "Good", "Damaged"]:
                messagebox.showerror("Error", "Invalid condition selected.")
                return False
            return self.db_manager.borrow_book_teacher(teacher_id, book_id, date.today(), condition)
        except TypeError as type_err:
            messagebox.showerror("Error", f"Invalid input type for borrowing: {type_err}")
            return False

    def return_book_teacher(self, teacher_id, book_id, condition=None):
        """Handle a teacher returning a book."""
        try:
            if not all([teacher_id, book_id]):
                messagebox.showerror("Error", "Please fill in all fields.")
                return False
            if condition and condition not in ["New", "Good", "Damaged"]:
                messagebox.showerror("Error", "Invalid condition selected.")
                return False
            return self.db_manager.return_book_teacher(teacher_id, book_id, condition)
        except TypeError as type_err:
            messagebox.showerror("Error", f"Invalid input type for returning: {type_err}")
            return False

    def view_all_books(self, tag_filter=None):
        """Fetch all books for display with optional tag filter."""
        try:
            books = self.db_manager.get_books_by_tags(tag_filter) if tag_filter else self.db_manager.load_books()
            if not books:
                messagebox.showinfo("Books", "No books available in the library.")
            return books
        except SQLiteError as db_err:
            messagebox.showerror("Database Error", f"Failed to load books: {db_err}")
            return []

    def get_all_books(self, tag_filter=None):
        """Fetch all books in the inventory with optional tag filter."""
        try:
            return self.db_manager.get_books_by_tags(tag_filter) if tag_filter else self.db_manager.load_books()
        except SQLiteError as db_err:
            messagebox.showerror("Database Error", f"Failed to fetch all books: {db_err}")
            return []

    def print_books(self, books):
        """Write the book list to a file."""
        try:
            with open("book_list.txt", "w") as file:
                file.write("BOOKS IN THE LIBRARY:\n\n")
                for book_id in books:
                    file.write(f"ID: {book_id}\n")
            return True
        except (FileNotFoundError, PermissionError, IOError) as file_err:
            messagebox.showerror("File Error", f"Failed to write books to file: {file_err}")
            return False

    def display_student_info(self, student_id, name_filter, filter_by_borrowed, filter_by_reams, result_label):
        """Display student information based on filters."""
        self.logger.info(f"Searching student - ID: {student_id}, Name: {name_filter}, Borrowed: {filter_by_borrowed}, Reams: {filter_by_reams}")
        try:
            conn = self.db_manager._create_connection()
            if not conn:
                result_label.config(text="Database connection failed.")
                return

            cursor = conn.cursor()
            query = "SELECT student_id, name FROM students WHERE 1=1"
            params = []

            if student_id:
                query += " AND student_id = ?"
                params.append(student_id)
            if name_filter:
                query += " AND name LIKE ?"
                params.append(f"%{name_filter}%")
            if filter_by_borrowed:
                query += " AND EXISTS (SELECT 1 FROM borrowed_books_student bbs WHERE bbs.student_id = students.student_id)"
            if filter_by_reams:
                query += " AND EXISTS (SELECT 1 FROM ream_entries r WHERE r.student_id = students.student_id AND r.reams_count > 0)"

            cursor.execute(query, params)
            students = cursor.fetchall()

            if not students:
                result_label.config(text="No students found matching the criteria.")
                return

            result_text = "Students Found:\n"
            for student_id, _ in students:
                name, stream, borrowed_books, reams_count, locker_id, chair_id = self.db_manager.get_student_info(student_id, conn)
                if name is None:
                    continue
                result_text += f"Student ID: {student_id}\nName: {name}\n"
                if stream:
                    result_text += f"Stream: {stream}\n"
                result_text += (f"Number of Reams Brought: {reams_count}\n"
                               f"Locker ID: {locker_id or 'Not assigned'}\n"
                               f"Chair ID: {chair_id or 'Not assigned'}\n"
                               "Borrowed Books:\n")
                result_text += "\n".join(f" - Book ID: {book_id}" for book_id in borrowed_books) if borrowed_books else " - No borrowed books found."
                result_text += "\n\n"

            result_label.config(text=result_text.strip())
        except SQLiteError as db_err:
            result_label.config(text=f"Database error: {db_err}")
        finally:
            self.db_manager._close_connection(conn)

    def display_teacher_info(self, teacher_id, name_filter, filter_by_borrowed, result_label):
        """Display teacher information based on filters."""
        self.logger.info(f"Searching teacher - ID: {teacher_id}, Name: {name_filter}, Borrowed: {filter_by_borrowed}")
        try:
            conn = self.db_manager._create_connection()
            if not conn:
                result_label.config(text="Database connection failed.")
                return

            cursor = conn.cursor()
            query = "SELECT teacher_id, teacher_name FROM teachers WHERE 1=1"
            params = []

            if teacher_id:
                query += " AND teacher_id = ?"
                params.append(teacher_id)
            if name_filter:
                query += " AND teacher_name LIKE ?"
                params.append(f"%{name_filter}%")
            if filter_by_borrowed:
                query += " AND EXISTS (SELECT 1 FROM borrowed_books_teacher bbt WHERE bbt.teacher_id = teachers.teacher_id)"

            cursor.execute(query, params)
            teachers = cursor.fetchall()

            if not teachers:
                result_label.config(text="No teachers found matching the criteria.")
                return

            result_text = "Teachers Found:\n"
            for teacher_id, name in teachers:
                name, borrowed_books = self.db_manager.get_teacher_info(teacher_id)
                result_text += f"Teacher ID: {teacher_id}\nName: {name}\nBorrowed Book IDs:\n"
                result_text += "\n".join(f" - Book ID: {b[0]} (Condition: {b[1]})" for b in borrowed_books) if borrowed_books else " - No borrowed books found."
                result_text += "\n\n"

            result_label.config(text=result_text.strip())
        except SQLiteError as db_err:
            result_label.config(text=f"Database error: {db_err}")
        finally:
            self.db_manager._close_connection(conn)

    def graduate_students(self):
        """Graduate all students to the next form."""
        try:
            conn = self.db_manager._create_connection()
            if not conn:
                return False, "Database connection failed."

            cursor = conn.cursor()
            cursor.execute("SELECT student_id, stream FROM students")
            students = cursor.fetchall()

            if not students:
                return True, "No students found to graduate."

            updated_count = 0
            for student_id, stream in students:
                if not stream:
                    continue
                match = re.match(r"(\d+)([A-Za-z]+)", stream)
                if match and int(match.group(1)) < 4:
                    new_stream = f"{int(match.group(1)) + 1}{match.group(2)}"
                    cursor.execute("UPDATE students SET stream = ? WHERE student_id = ?", (new_stream, student_id))
                    updated_count += 1

            conn.commit()
            return True, (f"Successfully graduated {updated_count} students." if updated_count else 
                         "No students needed graduation (all at max form or no stream).")
        except SQLiteError as db_err:
            conn.rollback()
            return False, f"Database error: {db_err}"
        finally:
            self.db_manager._close_connection(conn)

    def undo_graduate_students(self):
        """Revert all students to their previous form."""
        try:
            conn = self.db_manager._create_connection()
            if not conn:
                return False, "Database connection failed."

            cursor = conn.cursor()
            cursor.execute("SELECT student_id, stream FROM students")
            students = cursor.fetchall()

            if not students:
                return True, "No students found to undo."

            updated_count = 0
            for student_id, stream in students:
                if not stream:
                    continue
                match = re.match(r"(\d+)([A-Za-z]+)", stream)
                if match and int(match.group(1)) > 1:
                    new_stream = f"{int(match.group(1)) - 1}{match.group(2)}"
                    cursor.execute("UPDATE students SET stream = ? WHERE student_id = ?", (new_stream, student_id))
                    updated_count += 1

            conn.commit()
            return True, (f"Successfully reverted {updated_count} students." if updated_count else 
                         "No students needed undo (all at minimum form or no stream).")
        except SQLiteError as db_err:
            conn.rollback()
            return False, f"Database error: {db_err}"
        finally:
            self.db_manager._close_connection(conn)

    def update_student_info(self, student_id, new_stream=None, new_locker=None, new_chair=None):
        """Update student information (stream, locker, chair)."""
        self.logger.info(f"Updating student {student_id}: stream={new_stream}, locker={new_locker}, chair={new_chair}")
        if not isinstance(student_id, str) or not student_id:
            return False, "Student ID must be a non-empty string."

        try:
            conn = self.db_manager._create_connection()
            if not conn:
                return False, "Database connection failed."

            cursor = conn.cursor()
            cursor.execute("SELECT COUNT(*) FROM students WHERE student_id = ?", (student_id,))
            if cursor.fetchone()[0] == 0:
                return False, f"Student {student_id} not found."

            updates = []
            params = []
            if new_stream:
                if not isinstance(new_stream, str) or not re.match(r"^[1-4] [A-Za-z]+$", new_stream):
                    return False, f"Invalid stream format: {new_stream}"
                updates.append("stream = ?")
                params.append(new_stream)
            if new_locker:
                if not isinstance(new_locker, str):
                    return False, f"Locker ID must be a string, got {type(new_locker)}."
                cursor.execute("SELECT COUNT(*) FROM lockers WHERE locker_id = ? AND assigned = 0", (new_locker,))
                if cursor.fetchone()[0] == 0:
                    return False, f"Locker {new_locker} is invalid or already assigned."
            if new_chair:
                if not isinstance(new_chair, str):
                    return False, f"Chair ID must be a string, got {type(new_chair)}."
                cursor.execute("SELECT COUNT(*) FROM chairs WHERE chair_id = ? AND assigned = 0", (new_chair,))
                if cursor.fetchone()[0] == 0:
                    return False, f"Chair {new_chair} is invalid or already assigned."

            if not updates and not new_locker and not new_chair:
                return True, "No changes specified for update."

            if updates:
                # Build the SET clause safely with parameterized queries
                set_clause = ', '.join(updates)  # updates already contains '?' placeholders
                sql = f"UPDATE students SET {set_clause} WHERE student_id = ?"
                cursor.execute(sql, params + [student_id])
            if new_locker:
                cursor.execute("INSERT OR REPLACE INTO locker_assignments (student_id, locker_id) VALUES (?, ?)", 
                              (student_id, new_locker))
                cursor.execute("UPDATE lockers SET assigned = 1 WHERE locker_id = ?", (new_locker,))
            if new_chair:
                cursor.execute("INSERT OR REPLACE INTO chair_assignments (student_id, chair_id) VALUES (?, ?)", 
                              (student_id, new_chair))
                cursor.execute("UPDATE chairs SET assigned = 1 WHERE chair_id = ?", (new_chair,))

            conn.commit()
            updated_fields = [f for f, v in [("stream", new_stream), ("locker", new_locker), ("chair", new_chair)] if v]
            return True, f"Student {student_id} updated successfully: {', '.join(updated_fields)}."
        except SQLiteError as db_err:
            conn.rollback()
            return False, f"Database error: {db_err}"
        finally:
            self.db_manager._close_connection(conn)

    def get_unreturned_books(self):
        """Fetch unreturned books for display."""
        try:
            return self.db_manager.get_unreturned_books()
        except SQLiteError as db_err:
            messagebox.showerror("Database Error", f"Failed to fetch unreturned books: {db_err}")
            return {"students": [], "teachers": []}

    def add_ream(self, student_id, reams_brought):
        """Add reams for a student."""
        try:
            reams_brought = int(reams_brought)
            if reams_brought < 0:
                messagebox.showerror("Error", "Number of reams must be non-negative.")
                return False
            if self.db_manager.add_ream(student_id, reams_brought):
                messagebox.showinfo("Success", "Reams updated successfully!")
                return True
            return False
        except ValueError as val_err:
            messagebox.showerror("Error", f"Please enter a valid number for reams: {val_err}")
            return False
        except SQLiteError as db_err:
            messagebox.showerror("Database Error", f"Failed to add reams: {db_err}")
            return False

    def get_total_reams(self):
        """Fetch total reams and related data."""
        try:
            total_reams, student_reams, form_reams, form_stream_students, exported_reams = self.db_manager.get_total_reams()
            return total_reams, student_reams, form_reams, form_stream_students, True
        except SQLiteError as db_err:
            messagebox.showerror("Database Error", f"Failed to fetch reams: {db_err}")
            return 0, [], {}, {}, False

    def deduct_reams(self, used_reams, reams_remaining_label):
        """Deduct reams and update label."""
        try:
            used_reams = int(used_reams)
            if used_reams <= 0 or used_reams > 10000:
                messagebox.showerror("Error", "Number of reams must be positive and less than 10000.")
                return False
            
            # Debug: Check total reams before deduction
            total_reams, *_ = self.db_manager.get_total_reams()
            #print(f"Total reams before deduction: {total_reams}")
            
            if self.db_manager.deduct_reams(used_reams):
                total_reams, *_ = self.db_manager.get_total_reams()
                reams_remaining_label.config(text=f"Reams Remaining: {total_reams}")
                messagebox.showinfo("Success", "Reams deducted successfully!")
                return True
            return False
        except ValueError as val_err:
            messagebox.showerror("Error", f"Please enter a valid number of reams: {val_err}")
            return False
        except SQLiteError as db_err:
            messagebox.showerror("Database Error", f"Failed to deduct reams: {db_err}")
            return False

    def update_reams_remaining(self, label):
        """Update the reams remaining label."""
        try:
            total_reams, *_ = self.db_manager.get_total_reams()
            label.config(text=f"Reams Remaining: {total_reams}")
        except SQLiteError as db_err:
            messagebox.showerror("Database Error", f"Failed to update reams remaining: {db_err}")


    def get_all_reams_report(self, student_id_filter=None, date_start=None, date_end=None, min_reams=None, view_mode="detailed"):
        """Fetch and prepare the all reams report."""
        try:
            reams = self.db_manager.get_all_reams(student_id_filter, date_start, date_end, min_reams, view_mode)
            if not reams:
                messagebox.showinfo("Reams Report", f"No reams data matches the filters in {view_mode} view.")
            return reams
        except SQLiteError as db_err:
            messagebox.showerror("Error", f"Failed to fetch reams report: {db_err}")
            return []

    def download_all_reams_report(self, reams, output_dir="."):
        """Download the all reams report as a PDF."""
        try:
            pdf_filename = os.path.join(output_dir, f"All_Reams_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf")
            c = canvas.Canvas(pdf_filename, pagesize=letter)
            width, height = letter
            y_position = height - 50
            x_position = 50

            c.drawString(x_position, y_position, "All Reams Report")
            y_position -= 20
            c.drawString(x_position, y_position, f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
            y_position -= 20
            c.drawString(x_position, y_position, f"Total Records: {len(reams)}")
            y_position -= 40

            headers = ["Student ID", "Name", "Number of Reams", "Date Added"]
            col_widths = [100, 200, 100, 100]
            for i, header in enumerate(headers):
                c.drawString(x_position + sum(col_widths[:i]), y_position, header)
            y_position -= 20

            for student_id, name, reams_count, date_added in reams:
                c.drawString(x_position, y_position, str(student_id))
                c.drawString(x_position + col_widths[0], y_position, name[:30])
                c.drawString(x_position + col_widths[0] + col_widths[1], y_position, str(reams_count))
                c.drawString(x_position + col_widths[0] + col_widths[1] + col_widths[2], y_position, 
                            date_added.strftime('%Y-%m-%d') if isinstance(date_added, date) else str(date_added))
                y_position -= 20
                if y_position < 50:
                    c.showPage()
                    y_position = height - 50

            c.save()
            self.logger.info(f"All reams report downloaded as '{pdf_filename}'")
            return pdf_filename
        except Exception as e:
            messagebox.showerror("Error", f"Failed to download reams report: {e}")
            return None

    def export_reams_to_excel(self, output_dir):
        """Export ream entries to an Excel file."""
        try:
            conn = self.db_manager._create_connection()
            if not conn:
                messagebox.showerror("Error", "Database connection failed!")
                return False

            cursor = conn.cursor()
            cursor.execute("SELECT student_id, reams_count FROM ream_entries ORDER BY date_added DESC")
            rows = cursor.fetchall()

            if not rows:
                messagebox.showinfo("Info", "No ream entries found to export!")
                return False

            df = pd.DataFrame(rows, columns=['Student ID', 'Reams Brought'])
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"ream_entries_{timestamp}.xlsx"
            full_path = os.path.join(output_dir, filename)
            df.to_excel(full_path, index=False, engine='openpyxl')

            messagebox.showinfo("Success", f"Reams exported to:\n{full_path}")
            return True
        except Exception as e:
            messagebox.showerror("Error", f"Failed to export reams: {e}")
            return False
        finally:
            self.db_manager._close_connection(conn)

    def import_reams_from_excel(self, excel_file_path):
        """Import ream entries from an Excel file."""
        self.logger.info(f"Importing reams from {excel_file_path}")
        if not os.path.exists(excel_file_path):
            messagebox.showerror("Error", f"Excel file not found:\n{excel_file_path}")
            return False

        try:
            df = pd.read_excel(excel_file_path, engine='openpyxl')
            required_columns = ['Student ID', 'Reams Brought']
            if not all(col in df.columns for col in required_columns):
                messagebox.showerror("Error", f"Missing required columns: {set(required_columns) - set(df.columns)}")
                return False

            df = df.dropna(subset=required_columns)
            if df.empty:
                messagebox.showinfo("Info", "No valid data found in Excel file!")
                return False

            df['Student ID'] = df['Student ID'].astype(str)
            df['Reams Brought'] = df['Reams Brought'].astype(int)

            conn = self.db_manager._create_connection()
            if not conn:
                messagebox.showerror("Error", "Database connection failed!")
                return False

            cursor = conn.cursor()
            cursor.execute("SELECT student_id FROM students")
            existing_students = set(row[0] for row in cursor.fetchall())
            new_students = set(df['Student ID']) - existing_students

            if new_students:
                cursor.executemany("INSERT OR IGNORE INTO students (student_id) VALUES (?)", 
                                 [(sid,) for sid in new_students])
                conn.commit()

            success_count = 0
            failed_entries = []
            for _, row in df.iterrows():
                student_id, reams_count = row['Student ID'], row['Reams Brought']
                try:
                    cursor.execute("INSERT INTO ream_entries (student_id, reams_count, date_added) VALUES (?, ?, DATE('now'))",
                                  (student_id, reams_count))
                    success_count += 1
                except SQLiteError as e:
                    failed_entries.append((student_id, reams_count, str(e)))

            conn.commit()
            feedback = f"Imported {success_count} ream entries."
            if failed_entries:
                feedback += f"\nFailed {len(failed_entries)} entries:\n" + \
                           "\n".join(f"Student ID: {sid}, Reams: {r}, Error: {e}" for sid, r, e in failed_entries)
            messagebox.showinfo("Import Summary", feedback)
            return success_count > 0
        except (pd.errors.EmptyDataError, pd.errors.ParserError, ValueError) as e:
            messagebox.showerror("Error", f"Invalid Excel file or data: {e}")
            return False
        except SQLiteError as db_err:
            messagebox.showerror("Error", f"Failed to import reams: {db_err}")
            return False
        finally:
            self.db_manager._close_connection(conn)


    def search_student_for_return(self, student_id, student_name_label, book_list_frame, search_window=None):
        """Search for a student and display their borrowed books for return."""
        try:
            if not student_id:
                student_name_label.config(text="Please enter a student ID.")
                return

            student_info = self.db_manager.get_student_info(student_id)
            if student_info and student_info[0]:
                name, stream, borrowed_books = student_info[0], student_info[1], student_info[2]
                self.current_student_id = student_id

                display_text = f"Student Name: {name}"
                if stream:
                    display_text += f" | Stream: {stream}"
                student_name_label.config(text=display_text)

                if borrowed_books:
                    self._display_borrowed_books(borrowed_books, book_list_frame)
                else:
                    tk.Label(book_list_frame, text="No books borrowed").grid(row=0, column=0, columnspan=2)

                if search_window and isinstance(search_window, tk.Toplevel) and search_window.winfo_exists():
                    tk.Button(search_window, text="OK", command=search_window.destroy).grid(row=5, column=0, columnspan=2, padx=5, pady=5)
            else:
                student_name_label.config(text="Student Not Found")
                self._clear_book_list(book_list_frame)

        except SQLiteError as db_err:
            messagebox.showerror("Database Error", f"Failed to search student for return: {db_err}")
            student_name_label.config(text="Database error occurred.")
        except tk.TclError as tcl_err:
            messagebox.showerror("GUI Error", f"Failed to manipulate search window: {tcl_err}")

    def return_selected_books(self, student_id, book_list_frame, condition=None):
        """Return selected books for a student with the specified condition."""
        try:
            selected_books = [self.current_borrowed_books[i] for i, check_var in enumerate(self.check_vars) if check_var.get() == 1]

            if not selected_books:
                messagebox.showinfo("Information", "No books selected to return.")
                return False

            for book_id in selected_books:
                self.db_manager.return_book_student(student_id, book_id, condition)
            messagebox.showinfo("Success", "Selected books returned successfully!")

            self._clear_book_list(book_list_frame)
            student_info = self.db_manager.get_student_info(student_id)
            if student_info and student_info[0]:
                self._display_borrowed_books(student_info[2], book_list_frame)
            return True
        except (AttributeError, IndexError) as attr_err:
            messagebox.showerror("Error", f"Invalid state for returning books: {attr_err}")
            return False
        except tk.TclError as tcl_err:
            messagebox.showerror("GUI Error", f"Failed to refresh book list: {tcl_err}")
            return False

    def _display_borrowed_books(self, borrowed_books, book_list_frame):
        """Display borrowed books with checkboxes."""
        try:
            self._clear_book_list(book_list_frame)
            self.current_borrowed_books = borrowed_books
            self.check_vars = []
            for i, book_id in enumerate(borrowed_books):
                check_var = tk.IntVar()
                self.check_vars.append(check_var)
                tk.Checkbutton(book_list_frame, variable=check_var).grid(row=i, column=0, padx=5, pady=2)
                tk.Label(book_list_frame, text=f"Book ID: {book_id} (Condition: To be assessed)").grid(row=i, column=1, padx=5, pady=2)
        except AttributeError as attr_err:
            messagebox.showerror("Error", f"Failed to display borrowed books: {attr_err}")

    def _clear_book_list(self, book_list_frame):
        """Clear the book list frame."""
        try:
            for widget in book_list_frame.winfo_children():
                widget.destroy()
        except AttributeError as attr_err:
            messagebox.showerror("Error", f"Failed to clear book list: {attr_err}")

    def search_student_for_borrow(self, student_id, borrow_window, search_window, book_listbox, student_name_label):
        """Search a student for borrowing and show the borrow window."""
        try:
            if not student_id:
                messagebox.showerror("Error", "Please enter a student ID.")
                return
            student_info = self.db_manager.get_student_info(student_id)

            if student_info and student_info[0]:
                name, stream = student_info[0], student_info[1]
                self.current_student_id = student_id
                borrow_window.student_id = student_id

                display_text = f"Student Name: {name}"
                if stream:
                    display_text += f" | Stream: {stream}"
                student_name_label.config(text=display_text)

                search_window.destroy()
                borrow_window.deiconify()
                self.borrowed_books.clear()
                self._update_borrow_list(book_listbox)
            else:
                messagebox.showerror("Error", "Student not found.")
        except SQLiteError as db_err:
            messagebox.showerror("Database Error", f"Failed to search student for borrowing: {db_err}")
        except AttributeError as attr_err:
            messagebox.showerror("Error", f"Failed to update borrow window: {attr_err}")

    def add_book_to_borrow(self, book_id, book_listbox):
        try:
            if not book_id:
                messagebox.showerror("Error", "Please enter a book ID.")
                return
            self.borrowed_books.append(book_id)
            #print(f"After adding {book_id}: {self.borrowed_books}")  # Debug
            self._update_borrow_list(book_listbox)
        except AttributeError as attr_err:
            messagebox.showerror("Error", f"Failed to add book to borrow list: {attr_err}")

    def borrow_all_books(self, borrow_window, book_listbox):
        try:
            if not self.borrowed_books:
                messagebox.showerror("Error", "No books added to borrow.", parent=borrow_window)
                return False
            if not hasattr(self, 'current_student_id'):
                messagebox.showerror("Error", "No student selected.", parent=borrow_window)
                return False

            success, feedback = self.db_manager.borrow_book_student(
                self.current_student_id, self.borrowed_books[:], date.today(), reminder_days=None, condition="New"
            )

            if success:
                messagebox.showinfo("Borrowing Result", feedback, parent=borrow_window)
                self.borrowed_books.clear()
                self._update_borrow_list(book_listbox)  # Ensure listbox reflects cleared state
                #print(f"After borrow_all: {self.borrowed_books}")  # Debug
                return True
            else:
                messagebox.showerror("Borrowing Failed", feedback, parent=borrow_window)
                return False
        except SQLiteError as db_err:
            messagebox.showerror("Database Error", f"Failed to borrow books: {db_err}", parent=borrow_window)
            return False
        except AttributeError as attr_err:
            messagebox.showerror("Error", f"Failed to process borrowing: {attr_err}", parent=borrow_window)
            return False

    def _update_borrow_list(self, book_listbox):
        """Update the borrowing listbox."""
        try:
            if not isinstance(book_listbox, tk.Listbox):
                raise AttributeError(f"Expected Listbox, got {type(book_listbox)}")
            book_listbox.delete(0, tk.END)
            for i, book_id in enumerate(self.borrowed_books):
                book_listbox.insert(tk.END, f"{i + 1}) Book ID: {book_id}")
        except AttributeError as attr_err:
            messagebox.showerror("Error", f"Failed to update borrow list: {attr_err}")

    # Revision Books Control Logic
    def save_revision_book(self, book_id, tags):
        """Save a revision book with tags."""
        try:
            if not book_id:
                messagebox.showerror("Error", "Please enter Book ID.")
                return False
            if tags and not isinstance(tags, list):
                messagebox.showerror("Error", "Tags must be a list of strings.")
                return False
            if tags and any(not isinstance(tag, str) for tag in tags):
                messagebox.showerror("Error", "All tags must be strings.")
                return False
            return self.db_manager.save_revision_book(book_id, tags)
        except TypeError as type_err:
            self.logger.error(f"Invalid input type for revision book {book_id}: {type_err}")
            messagebox.showerror("Error", f"Invalid input type: {type_err}")
            return False
        except Exception as e:
            self.logger.error(f"Error saving revision book {book_id}: {e}")
            return False

    def borrow_revision_book(self, student_id, book_id, reminder_days, condition):
        """Borrow a revision book for a student."""
        try:
            if not all([student_id, book_id, condition]):
                self.logger.error("Missing required fields")
                return False
            if condition not in ["New", "Good", "Damaged"]:
                self.logger.error(f"Invalid condition: {condition}")
                return False
            reminder_days = 7 if reminder_days is None else int(reminder_days)
            if reminder_days <= 0:
                self.logger.error("Reminder days must be positive")
                return False

            success, feedback = self.db_manager.borrow_book_student(student_id, [book_id], date.today(), reminder_days, condition)
            if success:
                self.logger.info(f"Book {book_id} borrowed by {student_id}: {feedback}")
                return True
            else:
                self.logger.error(f"Failed to borrow book {book_id}: {feedback}")
                return False
        except (ValueError, SQLiteError) as e:
            self.logger.error(f"Error borrowing book {book_id}: {e}")
            return False

    def return_revision_book(self, student_id, book_list_frame):
        """Retrieve borrowed revision books for a student."""
        self.logger.debug(f"Retrieving revision books for student_id: {student_id}")
        try:
            if not student_id:
                self.logger.warning("No student ID provided")
                return None

            student_name = self.db_manager.get_student_name(student_id)
            if not student_name:
                self.logger.info(f"Student ID {student_id} not found")
                return None

            borrowed_books = self.db_manager.get_student_borrowed_revision_books(student_id)
            return student_name, borrowed_books or []
        except SQLiteError as db_err:
            self.logger.error(f"Database error retrieving revision books: {db_err}")
            return None

    def search_student_for_revision_return(self, student_id, student_name_label, book_list_frame):
        """Search a student and display borrowed revision books for return."""
        try:
            if not student_id:
                student_name_label.config(text="Please enter a student ID.")
                return

            name = self.db_manager.get_student_name(student_id)
            if name:
                student_name_label.config(text=f"Student Name: {name}")
                borrowed_books = self.db_manager.get_student_borrowed_revision_books(student_id)
                self._display_borrowed_revision_books(borrowed_books, book_list_frame)
            else:
                student_name_label.config(text="Student Not Found")
                self._clear_revision_book_list(book_list_frame)
        except SQLiteError as db_err:
            messagebox.showerror("Database Error", f"Failed to search student for return: {db_err}")
            student_name_label.config(text="Database error occurred.")

    def _display_borrowed_revision_books(self, borrowed_books, book_list_frame):
        """Display borrowed revision books in the frame."""
        self._clear_revision_book_list(book_list_frame)
        if borrowed_books:
            for book_id, borrowed_on, reminder_days in borrowed_books:
                book_info = f"Book ID: {book_id}, Borrowed On: {borrowed_on}, Reminder Days: {reminder_days}"
                check_var = tk.IntVar()
                check_button = tk.Checkbutton(book_list_frame, text=book_info, variable=check_var)
                check_button.var = check_var
                check_button.pack(anchor=tk.W)
        else:
            tk.Label(book_list_frame, text="No borrowed revision books found.").pack()

    def _clear_revision_book_list(self, book_list_frame):
        """Clear all widgets from the revision book list frame."""
        for widget in book_list_frame.winfo_children():
            widget.destroy()

    def return_selected_revision_books(self, student_id, book_list_frame, condition=None):
        """Return selected revision books for a student."""
        try:
            selected_books = [
                widget.cget("text").split("Book ID: ")[1].split(",")[0]
                for widget in book_list_frame.winfo_children()
                if isinstance(widget, tk.Checkbutton) and widget.var.get() == 1
            ]

            if not selected_books:
                messagebox.showinfo("Information", "No books selected to return.")
                return False

            for book_id in selected_books:
                self.db_manager.return_revision_book(student_id, book_id, condition)
            messagebox.showinfo("Success", "Selected revision books returned successfully!")

            if book_list_frame.winfo_children():
                self.search_student_for_revision_return(student_id, book_list_frame.master.winfo_children()[2], book_list_frame)
            return True
        except SQLiteError as db_err:
            messagebox.showerror("Database Error", f"Failed to return selected books: {db_err}")
            return False
        except Exception as e:
            messagebox.showerror("Error", f"An unexpected error occurred: {e}")
            return False

    def view_revision_books(self, tag_filter=None):
        """Fetch revision books with an optional tag filter."""
        try:
            books = self.db_manager.get_revision_books_by_tags(tag_filter) if tag_filter else self.db_manager.load_revision_books()
            if not books:
                messagebox.showinfo("Books", "No revision books available.")
            return books
        except SQLiteError as db_err:
            messagebox.showerror("Database Error", f"Failed to load revision books: {db_err}")
            return []

    def view_students_with_revision_books(self, student_id_filter=None, name_filter=None):
        """Fetch students with revision books with optional filters."""
        try:
            students = self.db_manager.get_students_with_revision_books(student_id_filter, name_filter)
            if not students:
                messagebox.showinfo("Students", "No students with revision books found.")
            return students
        except SQLiteError as db_err:
            messagebox.showerror("Database Error", f"Failed to fetch students with revision books: {db_err}")
            return []

    def save_reminder_settings(self, user_id, reminder_frequency, sound_enabled):
        """Save reminder settings for a user."""
        try:
            if reminder_frequency not in ["daily", "weekly", "disabled"]:
                messagebox.showerror("Error", "Invalid reminder frequency.")
                return False
            return self.db_manager.save_reminder_settings(user_id, reminder_frequency, sound_enabled)
        except SQLiteError as e:
            self.logger.error(f"Failed to save reminder settings: {e}")
            messagebox.showerror("Error", f"Failed to save reminder settings: {e}")
            return False

    def get_reminder_settings(self, user_id):
        """Fetch reminder settings for a user."""
        try:
            return self.db_manager.get_reminder_settings(user_id)
        except SQLiteError as e:
            self.logger.error(f"Failed to fetch reminder settings: {e}")
            messagebox.showerror("Error", f"Failed to fetch reminder settings: {e}")
            return "daily", True

    def get_books_by_condition_report(self, condition_filter=None):
        """Fetch and format a report of books by condition."""
        try:
            books_by_condition = self.db_manager.get_books_by_condition(condition_filter)
            if not any(books_by_condition.values()):
                messagebox.showinfo("Books by Condition", "No books found in the library.")
            return books_by_condition
        except SQLiteError as db_err:
            messagebox.showerror("Database Error", f"Failed to fetch books by condition: {db_err}")
            return {"New": [], "Good": [], "Damaged": []}

    def generate_qr_code(self, entity_type, entity_id):
        """Generate a QR code for an entity."""
        try:
            if not entity_id:
                messagebox.showerror("Error", f"{entity_type} ID cannot be empty.")
                return None, None
            if entity_type not in ["Book", "Student", "Teacher"]:
                messagebox.showerror("Error", "Invalid entity type selected.")
                return None, None

            qr = qrcode.QRCode(version=1, error_correction=qrcode.constants.ERROR_CORRECT_L, box_size=10, border=4)
            qr.add_data(f"{entity_type}:{entity_id}")
            qr.make(fit=True)
            qr_image = qr.make_image(fill_color="black", back_color="white")
            prefix = {"Book": "book", "Student": "student", "Teacher": "teacher"}
            filename = f"qr_{prefix[entity_type]}_{entity_id}.png"
            return filename, qr_image
        except Exception as e:
            self.logger.error(f"Failed to generate QR code for {entity_type} ID {entity_id}: {e}")
            messagebox.showerror("Error", f"Failed to generate QR code: {e}")
            return None, None

    def generate_bulk_book_qrcodes_pdf(self, book_class, subject, form, number_of_books, year, output_dir=".", save_directly=True):
        """Generate QR codes for multiple books and save to PDF or return images, 20 per A4 page."""
        try:
            number_of_books = int(number_of_books)
            if number_of_books <= 0:
                messagebox.showerror("Error", "Number of books must be greater than zero.")
                return None

            short_year = str(year)[-2:]
            pdf_filename = os.path.join(output_dir, f"Book_QRCodes_{book_class}_{subject}_{form}_{year}.pdf")
            pdf_buffer = io.BytesIO()
            c = canvas.Canvas(pdf_buffer, pagesize=letter)
            width, height = letter  # 612 x 792 points

            # Layout: 5 rows x 4 columns = 20 QR codes per page
            qr_size = 100  # QR code width/height in points
            text_height = 20  # Space for book_info text
            margin_x = 50  # Left/right margin
            margin_y = 50  # Top/bottom margin
            spacing_x = (width - 2 * margin_x - 4 * qr_size) / 3  # ~37 points
            spacing_y = (height - 2 * margin_y - 5 * (qr_size + text_height)) / 4  # ~23 points

            temp_files = []
            books_processed = 0

            for i in range(1, number_of_books + 1):
                book_info = f"{book_class}/{subject}/{form}/{i}/{short_year}"
                qr = qrcode.QRCode(version=1, error_correction=qrcode.constants.ERROR_CORRECT_L, box_size=10, border=4)
                qr.add_data(f"Book:{book_info}")
                qr.make(fit=True)
                qr_image = qr.make_image(fill_color="black", back_color="white")
                qr_image_filename = os.path.join(output_dir, f"temp_qrcode_{i}.png")
                qr_image.save(qr_image_filename)
                temp_files.append(qr_image_filename)

                # Calculate position in 5x4 grid
                page_book_index = books_processed % 20  # 0 to 19 for each page
                row = page_book_index // 4  # 0 to 4
                col = page_book_index % 4   # 0 to 3

                # Position from top-down
                x_position = margin_x + col * (qr_size + spacing_x)
                y_position = height - margin_y - row * (qr_size + text_height + spacing_y) - text_height

                # Draw text and QR code
                c.drawString(x_position, y_position, book_info)
                c.drawImage(qr_image_filename, x_position, y_position - qr_size, width=qr_size, height=qr_size)

                books_processed += 1
                if books_processed % 20 == 0 and books_processed < number_of_books:
                    c.showPage()  # New page after 20 books

            if books_processed % 20 != 0 or books_processed == 0:
                c.showPage()  # Save the last page

            c.save()
            pdf_data = pdf_buffer.getvalue()
            pdf_buffer.close()

            # Clean up temporary files
            for temp_file in temp_files:
                if os.path.exists(temp_file):
                    os.remove(temp_file)

            if save_directly:
                with open(pdf_filename, "wb") as f:
                    f.write(pdf_data)
                self.logger.info(f"Bulk QR codes saved to {pdf_filename}")
                return pdf_filename
            else:
                # Use PyMuPDF to convert PDF to images
                doc = fitz.open("pdf", pdf_data)
                images = []
                for page in doc:
                    pix = page.get_pixmap()
                    img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
                    images.append(img)
                doc.close()
                self.logger.info(f"Generated {len(images)} preview images for QR codes")
                return images

        except (ValueError, IOError) as e:
            self.logger.error(f"Failed to generate QR codes: {e}")
            messagebox.showerror("Error", f"Failed to generate QR codes: {e}")
            return None
        except Exception as e:
            self.logger.error(f"Unexpected error in QR code generation: {e}")
            messagebox.showerror("Error", f"Unexpected error: {e}")
            return None

    def export_library(self, output_dir=".", format="csv", progress_callback=None):
        """Export library data to CSV or JSON files."""
        try:
            data = self.db_manager.export_data()
            if not data or not any(data.values()):
                messagebox.showinfo("Export Info", "No data available to export.")
                return False

            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            total_steps = len(data)
            current_step = 0

            if format.lower() == "csv":
                for table, rows in data.items():
                    if not rows:
                        continue
                    df = pd.DataFrame(rows)
                    file_path = os.path.join(output_dir, f"{table}_{timestamp}.csv")
                    df.to_csv(file_path, index=False)
                    current_step += 1
                    if progress_callback:
                        progress_callback(current_step / total_steps * 100)
                messagebox.showinfo("Success", f"Library data exported as CSV files in {output_dir}")
            elif format.lower() == "json":
                file_path = os.path.join(output_dir, f"library_data_{timestamp}.json")
                with open(file_path, "w") as f:
                    json.dump(data, f, indent=4, default=str)
                if progress_callback:
                    progress_callback(100)
                messagebox.showinfo("Success", f"Library data exported as JSON to {file_path}")
            else:
                messagebox.showerror("Export Error", "Unsupported format. Use 'csv' or 'json'.")
                return False
            return True
        except Exception as e:
            messagebox.showerror("Export Error", f"Failed to export library data: {e}")
            return False

    def import_library(self, file_path=None, directory=None, format="csv", incremental=False, progress_callback=None):
        """Import library data from CSV or JSON files."""
        try:
            if format.lower() == "csv" and directory:
                data = {}
                files = [f for f in os.listdir(directory) if f.endswith(".csv")]
                if not files:
                    messagebox.showinfo("Import Info", "No CSV files found in the directory.")
                    return False
                total_steps = len(files)
                current_step = 0
                for file in files:
                    table = file.split('_')[0]
                    df = pd.read_csv(os.path.join(directory, file))
                    data[table] = df.to_dict(orient="records")
                    current_step += 1
                    if progress_callback:
                        progress_callback(current_step / total_steps * 100)
            elif format.lower() == "json" and file_path:
                with open(file_path, "r") as f:
                    data = json.load(f)
                if not data or not any(data.values()):
                    messagebox.showinfo("Import Info", "No data found in JSON file.")
                    return False
                if progress_callback:
                    progress_callback(100)
            else:
                messagebox.showerror("Import Error", "Invalid format or missing file/directory.")
                return False

            import_method = self.db_manager.incremental_import_data if incremental else self.db_manager.import_data
            if import_method(data):
                mode = "incrementally" if incremental else "fully"
                messagebox.showinfo("Success", f"Library data imported {mode} successfully!")
                return True
            return False
        except (FileNotFoundError, json.JSONDecodeError) as e:
            messagebox.showerror("Import Error", f"Invalid file or data: {e}")
            return False
        except Exception as e:
            messagebox.showerror("Import Error", f"Failed to import library data: {e}")
            return False

    
    def delete_from_system(self, entity_type, identifier):
        """Delete an entity from the system."""
        conn = self.db_manager._create_connection()
        if not conn:
            self.logger.error("Failed to connect to database for deletion")
            return False

        cursor = conn.cursor()
        try:
            queries = {
                "student": "DELETE FROM students WHERE student_id = ?",
                "book": "DELETE FROM books WHERE book_id = ?",
                "borrowed_book_student": "DELETE FROM borrowed_books_student WHERE student_id = ? AND book_id = ?",
                "teacher": "DELETE FROM teachers WHERE teacher_id = ?",
                "borrowed_book_teacher": "DELETE FROM borrowed_books_teacher WHERE teacher_id = ? AND book_id = ?",
                "ream_entry": "DELETE FROM ream_entries WHERE id = ?"
            }
            if entity_type not in queries:
                raise ValueError(f"Unknown entity type: {entity_type}")

            params = identifier if isinstance(identifier, tuple) else (identifier,)
            cursor.execute(queries[entity_type], params)
            conn.commit()
            self.logger.info(f"Deleted {entity_type} with identifier {identifier}")
            return True
        except (SQLiteError, ValueError) as e:
            conn.rollback()
            self.logger.error(f"Error deleting {entity_type} with identifier {identifier}: {e}")
            return False
        finally:
            self.db_manager._close_connection(conn)

    def search_book_borrow_details(self, book_id):
        """Search for borrowing details of a book by its ID."""
        conn = self.db_manager._create_connection()
        if not conn:
            self.logger.error("Failed to connect to database for book search")
            return None

        cursor = conn.cursor()
        try:
            cursor.execute("""
                SELECT s.student_id, s.name, bbs.borrowed_on
                FROM borrowed_books_student bbs
                JOIN students s ON bbs.student_id = s.student_id
                WHERE bbs.book_id = ?
            """, (book_id,))
            student_result = cursor.fetchone()

            if student_result:
                return {
                    "book_id": book_id,
                    "borrower_type": "student",
                    "borrower_id": student_result[0],
                    "borrower_name": student_result[1],
                    "borrowed_on": student_result[2]
                }

            cursor.execute("""
                SELECT t.teacher_id, t.teacher_name, bbt.borrowed_on
                FROM borrowed_books_teacher bbt
                JOIN teachers t ON bbt.teacher_id = t.teacher_id
                WHERE bbt.book_id = ?
            """, (book_id,))
            teacher_result = cursor.fetchone()
            if teacher_result:
                return {
                    "book_id": book_id,
                    "borrower_type": "teacher",
                    "borrower_id": teacher_result[0],
                    "borrower_name": teacher_result[1],
                    "borrowed_on": teacher_result[2]
                }

            cursor.execute("SELECT book_id FROM books WHERE book_id = ?", (book_id,))
            if cursor.fetchone():
                return {"book_id": book_id, "borrower_type": "none", "message": "Book is not currently borrowed"}
            return None
        except SQLiteError as e:
            self.logger.error(f"Error searching book {book_id}: {e}")
            return None
        finally:
            self.db_manager._close_connection(conn)

    def generate_furniture_id(self, furniture_type, color, count=1, form=None):
        """Generate sequential IDs for lockers or chairs."""
        conn = self.db_manager._create_connection()
        if not conn:
            self.logger.error("Failed to connect to database for ID generation")
            return []

        cursor = conn.cursor()
        try:
            table = "lockers" if furniture_type == "LKR" else "chairs"
            id_field = "locker_id" if furniture_type == "LKR" else "chair_id"
            prefix = f"{furniture_type}/{color}/" if not form else f"{form}/{furniture_type}/{color}/"

            cursor.execute(f"SELECT {id_field} FROM {table} WHERE {id_field} LIKE ? ORDER BY {id_field} DESC LIMIT 1",
                          (f"{prefix}%",))
            last_id = cursor.fetchone()
            next_seq = int(last_id[0].split("/")[-1]) + 1 if last_id else 1

            return [f"{prefix}{str(next_seq + i).zfill(3)}" for i in range(count)]
        except SQLiteError as e:
            self.logger.error(f"Error generating furniture IDs: {e}")
            return []
        finally:
            self.db_manager._close_connection(conn)

    def validate_furniture_ids(self, furniture_type, ids):
        """Validate a list of furniture IDs."""
        if not ids or not isinstance(ids, list):
            return False, "IDs must be a non-empty list."

        conn = self.db_manager._create_connection()
        if not conn:
            return False, "Database connection failed."

        cursor = conn.cursor()
        try:
            table = "lockers" if furniture_type == "LKR" else "chairs"
            id_field = "locker_id" if furniture_type == "LKR" else "chair_id"
            for fid in ids:
                if not isinstance(fid, str) or not fid.strip() or len(fid) > 20:
                    return False, f"Invalid ID: '{fid}'. Must be a non-empty string under 20 characters."
                cursor.execute(f"SELECT COUNT(*) FROM {table} WHERE {id_field} = ?", (fid,))
                if cursor.fetchone()[0] > 0:
                    return False, f"ID '{fid}' already exists in {table}."
            return True, ""
        except SQLiteError as e:
            return False, f"Error validating IDs: {e}"
        finally:
            self.db_manager._close_connection(conn)

    def add_furniture(self, furniture_type, ids=None, location=None, form=None, color=None, cond="Good", count=1):
        """Add new lockers or chairs to the database with simple ID format (e.g., LKR/R/001)."""
        if furniture_type not in ["LKR", "CHR"]:
            return False, "Furniture type must be 'LKR' or 'CHR'."

        if ids is None:
            if not color:
                return False, "Color must be provided when IDs are not specified."
            # Generate simple IDs
            ids = [f"{furniture_type}/{color}/{i:03d}" for i in range(1, count + 1)]

        # Optional validation (you can enhance this if needed)
        valid, error = self.validate_furniture_ids(furniture_type, ids)
        if not valid:
            return False, error

        conn = self.db_manager._create_connection()
        if not conn:
            return False, "Database connection failed."

        cursor = conn.cursor()
        try:
            table = "lockers" if furniture_type == "LKR" else "chairs"
            id_field = "locker_id" if furniture_type == "LKR" else "chair_id"
            for fid in ids:
                cursor.execute(f"""
                    INSERT INTO {table} ({id_field}, location, form, color, cond)
                    VALUES (?, ?, ?, ?, ?)
                """, (fid, location, form, color, cond))
            conn.commit()
            self.logger.info(f"Added {len(ids)} {furniture_type} items")
            return True, ids
        except SQLiteError as e:
            conn.rollback()
            return False, f"Error adding furniture: {e}"
        finally:
            self.db_manager._close_connection(conn)

    def unassign_locker(self, student_id, locker_id):
        """Unassign a locker from a student."""
        conn = self.db_manager._create_connection()
        if not conn:
            self.logger.error("Failed to connect to database for unassigning locker")
            return False

        cursor = conn.cursor()
        try:
            cursor.execute("DELETE FROM locker_assignments WHERE student_id = ? AND locker_id = ?", (student_id, locker_id))
            cursor.execute("UPDATE lockers SET assigned = 0 WHERE locker_id = ?", (locker_id,))
            conn.commit()
            self.logger.info(f"Unassigned locker {locker_id} from student {student_id}")
            return True
        except SQLiteError as e:
            conn.rollback()
            self.logger.error(f"Error unassigning locker: {e}")
            return False
        finally:
            self.db_manager._close_connection(conn)

    def get_furniture_status(self, furniture_type):
        """Retrieve the status of furniture with assignment details, aligned with FurnitureAssignmentGUI."""
        conn = self.db_manager._create_connection()
        if not conn:
            self.logger.error(f"Failed to connect to database for {furniture_type} status")
            return []

        cursor = conn.cursor()
        try:
            table = "lockers" if furniture_type == "lockers" else "chairs"
            id_field = "locker_id" if furniture_type == "lockers" else "chair_id"
            assignment_table = "locker_assignments" if furniture_type == "lockers" else "chair_assignments"
            cursor.execute(f"""
                SELECT 
                    t.{id_field},
                    t.location,
                    t.form,
                    t.color,
                    t.cond,
                    t.assigned,
                    a.student_id,
                    s.name
                FROM {table} t
                LEFT JOIN {assignment_table} a ON t.{id_field} = a.{id_field}
                LEFT JOIN students s ON a.student_id = s.student_id
            """)
            results = [
                {
                    id_field: r[0],  # "locker_id" or "chair_id"
                    "location": r[1],
                    "form": r[2],
                    "color": r[3],
                    "cond": r[4],
                    "assigned": bool(r[5]),
                    "student_id": r[6],
                    "student_name": r[7]
                } for r in cursor.fetchall()
            ]
            self.logger.info(f"Retrieved status for {len(results)} {furniture_type}")
            return results
        except SQLiteError as e:
            self.logger.error(f"Error retrieving {furniture_type} status: {e}")
            return []
        finally:
            self.db_manager._close_connection(conn)

    def report_damaged_furniture(self, furniture_type, furniture_id, condition):
        """Report a piece of furniture as damaged."""
        valid_conditions = ["Good", "Damaged", "Needs Repair"]
        if furniture_type not in ["lockers", "chairs"] or not furniture_id or condition not in valid_conditions:
            self.logger.error(f"Invalid input: type={furniture_type}, id={furniture_id}, condition={condition}")
            return False

        conn = self.db_manager._create_connection()
        if not conn:
            self.logger.error(f"Failed to connect to database to report damaged {furniture_type}")
            return False

        cursor = conn.cursor()
        try:
            table = "lockers" if furniture_type == "lockers" else "chairs"
            id_field = "locker_id" if furniture_type == "lockers" else "chair_id"
            cursor.execute(f"UPDATE {table} SET cond = ? WHERE {id_field} = ?", (condition, furniture_id))
            if cursor.rowcount == 0:
                conn.rollback()
                return False
            conn.commit()
            self.logger.info(f"Reported {furniture_type} {furniture_id} as {condition}")
            return True
        except SQLiteError as e:
            conn.rollback()
            self.logger.error(f"Error reporting damaged {furniture_type}: {e}")
            return False
        finally:
            self.db_manager._close_connection(conn)



    def assign_locker(self, student_id, locker_id):
        """Assign a locker to a student."""
        if not student_id or not isinstance(student_id, str):
            self.logger.error(f"Invalid student ID: {student_id}. Must be a non-empty string")
            return False
        if not locker_id or not isinstance(locker_id, str) or len(locker_id) > 20:
            self.logger.error(f"Invalid locker ID: {locker_id}. Must be a non-empty string up to 20 characters")
            return False

        conn = self.db_manager._create_connection()
        if not conn:
            self.logger.error("Failed to connect to database for locker assignment")
            return False

        cursor = conn.cursor()
        try:
            cursor.execute("SELECT assigned FROM lockers WHERE locker_id = ?", (locker_id,))
            result = cursor.fetchone()
            if not result:
                self.logger.error(f"Locker {locker_id} does not exist")
                return False
            if result[0] == 1:
                self.logger.error(f"Locker {locker_id} is already assigned")
                return False

            cursor.execute("SELECT 1 FROM students WHERE student_id = ?", (student_id,))
            if not cursor.fetchone():
                self.logger.error(f"Student {student_id} does not exist")
                return False

            cursor.execute("UPDATE lockers SET assigned = 1 WHERE locker_id = ?", (locker_id,))
            cursor.execute("""
                INSERT OR REPLACE INTO locker_assignments (student_id, locker_id, assigned_date)
                VALUES (?, ?, DATE('now'))
            """, (student_id, locker_id))
            conn.commit()
            self.logger.info(f"Assigned locker {locker_id} to student {student_id}")
            return True
        except SQLiteError as e:
            self.logger.error(f"Error assigning locker: {e}")
            conn.rollback()
            return False
        finally:
            self.db_manager._close_connection(conn)

    def assign_chair(self, student_id, chair_id):
        """Assign a chair to a student."""
        if not student_id or not isinstance(student_id, str):
            self.logger.error(f"Invalid student ID: {student_id}. Must be a non-empty string")
            return False
        if not chair_id or not isinstance(chair_id, str) or len(chair_id) > 20:
            self.logger.error(f"Invalid chair ID: {chair_id}. Must be a non-empty string up to 20 characters")
            return False

        conn = self.db_manager._create_connection()
        if not conn:
            self.logger.error("Failed to connect to database for chair assignment")
            return False

        cursor = conn.cursor()
        try:
            cursor.execute("SELECT assigned FROM chairs WHERE chair_id = ?", (chair_id,))
            result = cursor.fetchone()
            if not result:
                self.logger.error(f"Chair {chair_id} does not exist")
                return False
            if result[0] == 1:
                self.logger.error(f"Chair {chair_id} is already assigned")
                return False

            cursor.execute("SELECT 1 FROM students WHERE student_id = ?", (student_id,))
            if not cursor.fetchone():
                self.logger.error(f"Student {student_id} does not exist")
                return False

            cursor.execute("UPDATE chairs SET assigned = 1 WHERE chair_id = ?", (chair_id,))
            cursor.execute("""
                INSERT OR REPLACE INTO chair_assignments (student_id, chair_id, assigned_date)
                VALUES (?, ?, DATE('now'))
            """, (student_id, chair_id))
            conn.commit()
            self.logger.info(f"Assigned chair {chair_id} to student {student_id}")
            return True
        except SQLiteError as e:
            self.logger.error(f"Error assigning chair: {e}")
            conn.rollback()
            return False
        finally:
            self.db_manager._close_connection(conn)

    def search_furniture_by_id(self, furniture_type, furniture_id):
        """Search for student details assigned to a specific locker or chair by ID."""
        if furniture_type not in ["lockers", "chairs"]:
            self.logger.error(f"Invalid furniture type: {furniture_type}. Must be 'lockers' or 'chairs'")
            return None
        if not furniture_id or not isinstance(furniture_id, str):
            self.logger.error(f"Invalid furniture ID: {furniture_id}. Must be a non-empty string")
            return None

        conn = self.db_manager._create_connection()
        if not conn:
            self.logger.error(f"Failed to connect to database for {furniture_type} search")
            return None

        cursor = conn.cursor()
        try:
            table = "lockers" if furniture_type == "lockers" else "chairs"
            id_field = "locker_id" if furniture_type == "lockers" else "chair_id"
            assignment_table = "locker_assignments" if furniture_type == "lockers" else "chair_assignments"
            query = f"""
                SELECT 
                    t.{id_field},
                    t.location,
                    t.form,
                    t.color,
                    t.cond,
                    t.assigned,
                    a.student_id,
                    s.name,
                    s.stream,
                    a.assigned_date
                FROM {table} t
                LEFT JOIN {assignment_table} a ON t.{id_field} = a.{id_field}
                LEFT JOIN students s ON a.student_id = s.student_id
                WHERE t.{id_field} = ?
            """
            cursor.execute(query, (furniture_id,))
            result = cursor.fetchone()
            if result:
                self.logger.info(f"Found {furniture_type} {furniture_id}: assigned to {result[6] or 'Unassigned'}")
                return {
                    id_field: result[0], "location": result[1], "form": result[2], "color": result[3],
                    "cond": result[4], "assigned": bool(result[5]), "student_id": result[6],
                    "name": result[7], "stream": result[8], "assigned_date": result[9]
                }
            self.logger.warning(f"No {furniture_type} found with ID {furniture_id}")
            return None
        except SQLiteError as e:
            self.logger.error(f"Error searching {furniture_type} by ID: {e}")
            return None
        finally:
            self.db_manager._close_connection(conn)

    def get_furniture_totals_by_form(self):
        """Retrieve total lockers and chairs, categorized by form."""
        conn = self.db_manager._create_connection()
        if not conn:
            self.logger.error("Failed to connect to database for furniture totals")
            return None

        cursor = conn.cursor()
        try:
            cursor.execute("SELECT COUNT(*) FROM lockers")
            total_lockers = cursor.fetchone()[0]
            cursor.execute("SELECT COUNT(*) FROM chairs")
            total_chairs = cursor.fetchone()[0]

            cursor.execute("SELECT form, COUNT(*) FROM lockers GROUP BY form")
            lockers_by_form = dict(cursor.fetchall())
            cursor.execute("SELECT form, COUNT(*) FROM chairs GROUP BY form")
            chairs_by_form = dict(cursor.fetchall())

            result = {
                "total_lockers": total_lockers,
                "total_chairs": total_chairs,
                "lockers_by_form": lockers_by_form,
                "chairs_by_form": chairs_by_form
            }
            self.logger.info(f"Retrieved furniture totals: {total_lockers} lockers, {total_chairs} chairs")
            return result
        except SQLiteError as e:
            self.logger.error(f"Error retrieving furniture totals: {e}")
            return None
        finally:
            self.db_manager._close_connection(conn)

    def get_all_furniture_details(self):
        """Retrieve detailed list of all furniture with assignment info."""
        conn = self.db_manager._create_connection()
        if not conn:
            self.logger.error("Failed to connect to database for furniture details")
            return None

        cursor = conn.cursor()
        try:
            cursor.execute("""
                SELECT 
                    l.locker_id,
                    l.form,
                    l.cond,
                    la.student_id,
                    s.name
                FROM lockers l
                LEFT JOIN locker_assignments la ON l.locker_id = la.locker_id
                LEFT JOIN students s ON la.student_id = s.student_id
            """)
            lockers = [{"locker_id": r[0], "form": r[1], "cond": r[2], "student_id": r[3], "name": r[4]} for r in cursor.fetchall()]

            cursor.execute("""
                SELECT 
                    c.chair_id,
                    c.form,
                    c.cond,
                    ca.student_id,
                    s.name
                FROM chairs c
                LEFT JOIN chair_assignments ca ON c.chair_id = ca.chair_id
                LEFT JOIN students s ON ca.student_id = s.student_id
            """)
            chairs = [{"chair_id": r[0], "form": r[1], "cond": r[2], "student_id": r[3], "name": r[4]} for r in cursor.fetchall()]

            result = {"lockers": lockers, "chairs": chairs}
            self.logger.info(f"Retrieved details for {len(lockers)} lockers and {len(chairs)} chairs")
            return result
        except SQLiteError as e:
            self.logger.error(f"Error retrieving furniture details: {e}")
            return None
        finally:
            self.db_manager._close_connection(conn)

    def add_or_update_furniture_category(self, category_name, total_count, needs_repair=0):
        """Add or update a furniture category with total count and repair needs."""
        if not category_name or not isinstance(category_name, str):
            self.logger.error(f"Invalid category name: {category_name}. Must be a non-empty string")
            return False
        if not isinstance(total_count, int) or total_count < 0:
            self.logger.error(f"Invalid total count: {total_count}. Must be a non-negative integer")
            return False
        if not isinstance(needs_repair, int) or needs_repair < 0 or needs_repair > total_count:
            self.logger.error(f"Invalid needs_repair: {needs_repair}. Must be 0 to {total_count}")
            return False

        conn = self.db_manager._create_connection()
        if not conn:
            self.logger.error("Failed to connect to database for furniture category update")
            return False

        cursor = conn.cursor()
        try:
            cursor.execute("""
                INSERT OR REPLACE INTO furniture_categories (category_name, total_count, needs_repair)
                VALUES (?, ?, ?)
            """, (category_name, total_count, needs_repair))
            conn.commit()
            self.logger.info(f"Updated category '{category_name}': {total_count} total, {needs_repair} need repair")
            return True
        except SQLiteError as e:
            self.logger.error(f"Error updating furniture category: {e}")
            conn.rollback()
            return False
        finally:
            self.db_manager._close_connection(conn)

    def get_all_furniture_categories(self):
        """Retrieve all furniture categories with their counts."""
        conn = self.db_manager._create_connection()
        if not conn:
            self.logger.error("Failed to connect to database for furniture categories")
            return []

        cursor = conn.cursor()
        try:
            cursor.execute("""
                SELECT category_name, total_count, needs_repair
                FROM furniture_categories
                ORDER BY category_name
            """)
            categories = [{"category_name": r[0], "total_count": r[1], "needs_repair": r[2]} for r in cursor.fetchall()]
            self.logger.info(f"Retrieved {len(categories)} furniture categories")
            return categories
        except SQLiteError as e:
            self.logger.error(f"Error retrieving furniture categories: {e}")
            return []
        finally:
            self.db_manager._close_connection(conn)

    def load_chairs_from_excel(self, file_path):
        """Load chair assignments from an Excel file into the database."""
        try:
            self.logger.info(f"Loading chairs from Excel file: {file_path}")
            df = pd.read_excel(file_path)
            required_columns = {'student_id', 'chair_id'}
            if not required_columns.issubset(df.columns):
                self.logger.error(f"Excel file missing required columns: {required_columns}")
                return False, f"Excel file must contain {required_columns}"

            conn = self.db_manager._create_connection()
            if not conn:
                self.logger.error("Database connection failed")
                return False, "Database connection failed."

            cursor = conn.cursor()
            assigned_count = 0

            for index, row in df.iterrows():
                student_id = str(row['student_id']).strip()
                chair_id = str(row['chair_id']).strip()
                if len(chair_id) > 20:
                    self.logger.warning(f"Chair ID {chair_id} exceeds 20 characters, skipping row {index}")
                    continue
                assigned_date = row.get('assigned_date', datetime.now().date())

                cursor.execute("SELECT COUNT(*) FROM students WHERE student_id = ?", (student_id,))
                if cursor.fetchone()[0] == 0:
                    self.logger.warning(f"Student {student_id} not found, skipping row {index}")
                    continue

                cursor.execute("SELECT assigned FROM chairs WHERE chair_id = ?", (chair_id,))
                chair_result = cursor.fetchone()
                if not chair_result:
                    self.logger.warning(f"Chair {chair_id} not found, skipping row {index}")
                    continue
                if chair_result[0] == 1:
                    self.logger.warning(f"Chair {chair_id} already assigned, skipping row {index}")
                    continue

                cursor.execute("""
                    INSERT OR REPLACE INTO chair_assignments (student_id, chair_id, assigned_date)
                    VALUES (?, ?, ?)
                """, (student_id, chair_id, assigned_date))
                cursor.execute("UPDATE chairs SET assigned = 1 WHERE chair_id = ?", (chair_id,))
                self.logger.info(f"Assigned chair {chair_id} to student {student_id}")
                assigned_count += 1

            conn.commit()
            self.logger.info(f"Successfully loaded {assigned_count} chair assignments")
            return True, f"Loaded {assigned_count} chair assignments from Excel."
        except (pd.errors.EmptyDataError, SQLiteError) as e:
            self.logger.error(f"Failed to load chairs: {e}")
            conn.rollback()
            return False, f"Error loading chairs: {e}"
        finally:
            self.db_manager._close_connection(conn)

    def load_lockers_from_excel(self, file_path):
        """Load locker assignments from an Excel file into the database."""
        try:
            self.logger.info(f"Loading lockers from Excel file: {file_path}")
            df = pd.read_excel(file_path)
            required_columns = {'student_id', 'locker_id'}
            if not required_columns.issubset(df.columns):
                self.logger.error(f"Excel file missing required columns: {required_columns}")
                return False, f"Excel file must contain {required_columns}"

            conn = self.db_manager._create_connection()
            if not conn:
                self.logger.error("Database connection failed")
                return False, "Database connection failed."

            cursor = conn.cursor()
            assigned_count = 0

            for index, row in df.iterrows():
                student_id = str(row['student_id']).strip()
                locker_id = str(row['locker_id']).strip()
                if len(locker_id) > 20:
                    self.logger.warning(f"Locker ID {locker_id} exceeds 20 characters, skipping row {index}")
                    continue
                assigned_date = row.get('assigned_date', datetime.now().date())

                cursor.execute("SELECT COUNT(*) FROM students WHERE student_id = ?", (student_id,))
                if cursor.fetchone()[0] == 0:
                    self.logger.warning(f"Student {student_id} not found, skipping row {index}")
                    continue

                cursor.execute("SELECT assigned FROM lockers WHERE locker_id = ?", (locker_id,))
                locker_result = cursor.fetchone()
                if not locker_result:
                    self.logger.warning(f"Locker {locker_id} not found, skipping row {index}")
                    continue
                if locker_result[0] == 1:
                    self.logger.warning(f"Locker {locker_id} already assigned, skipping row {index}")
                    continue

                cursor.execute("""
                    INSERT OR REPLACE INTO locker_assignments (student_id, locker_id, assigned_date)
                    VALUES (?, ?, ?)
                """, (student_id, locker_id, assigned_date))
                cursor.execute("UPDATE lockers SET assigned = 1 WHERE locker_id = ?", (locker_id,))
                self.logger.info(f"Assigned locker {locker_id} to student {student_id}")
                assigned_count += 1

            conn.commit()
            self.logger.info(f"Successfully loaded {assigned_count} locker assignments")
            return True, f"Loaded {assigned_count} locker assignments from Excel."
        except (pd.errors.EmptyDataError, SQLiteError) as e:
            self.logger.error(f"Failed to load lockers: {e}")
            conn.rollback()
            return False, f"Error loading lockers: {e}"
        finally:
            self.db_manager._close_connection(conn)

    def display_students_by_form(self):
        """Fetch all students, group them by form, and return a dictionary."""
        try:
            self.logger.info("Fetching students grouped by form")
            conn = self.db_manager._create_connection()
            if not conn:
                self.logger.error("Database connection failed")
                return False, "Database connection failed."

            cursor = conn.cursor()
            try:
                cursor.execute("SELECT student_id, name, stream FROM students ORDER BY student_id")
                students = cursor.fetchall()

                if not students:
                    self.logger.info("No students found in the database")
                    return True, {"forms": {}}

                form_groups = {f"Form {i}": [] for i in range(1, 5)}
                for student_id, name, stream in students:
                    if not stream:
                        continue
                    match = re.match(r"(\d+)", stream)
                    if match and 1 <= int(match.group(1)) <= 4:
                        form_key = f"Form {match.group(1)}"
                        form_groups[form_key].append({"student_id": student_id, "name": name, "stream": stream})

                result = {"forms": {}}
                for form, students_list in form_groups.items():
                    result["forms"][form] = {
                        "total": len(students_list),
                        "students": sorted(students_list, key=lambda x: x["student_id"])
                    }

                self.logger.info("Successfully grouped students by form")
                return True, result
            except SQLiteError as e:
                self.logger.error(f"Failed to fetch students by form: {e}")
                return False, f"Error fetching students: {e}"
            finally:
                self.db_manager._close_connection(conn)
        except Exception as e:
            self.logger.error(f"Unexpected error in display_students_by_form: {e}")
            return False, f"Unexpected error: {e}"

    def add_short_form_mapping(self, short_form, full_name, mapping_type):
        """Add a new short form mapping."""
        return self.db_manager.add_short_form_mapping(short_form, full_name, mapping_type)

    def get_short_form_mappings(self):
        """Fetch all short form mappings."""
        return self.db_manager.get_short_form_mappings()

    def get_students_without_borrowed_books(self):
        """Fetch students who have not borrowed any books, sorted by stream."""
        conn = self.db_manager._create_connection()
        if not conn:
            self.logger.error("Failed to connect to database for students without borrowed books")
            messagebox.showerror("Database Error", "Failed to connect to the database.")
            return []

        cursor = conn.cursor()
        try:
            query = """
                SELECT s.student_id, s.name, s.stream
                FROM students s
                LEFT JOIN borrowed_books_student bbs ON s.student_id = bbs.student_id
                WHERE bbs.student_id IS NULL
                ORDER BY s.stream
            """
            cursor.execute(query)
            students = cursor.fetchall()
            self.logger.info(f"Retrieved {len(students)} students without borrowed books")
            return students
        except SQLiteError as e:
            self.logger.error(f"Error retrieving students without borrowed books: {e}")
            messagebox.showerror("Database Error", f"Failed to fetch students without borrowed books: {e}")
            return []
        finally:
            self.db_manager._close_connection(conn)

